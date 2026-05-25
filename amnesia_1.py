import numpy as np
from openpyxl import load_workbook

FILE_PATH = "<filepath>"   
SHEET1 = "Sheet1"
SHEET2 = "Sheet2"

# Sheet1 layout
S1_START_COL = 2
S1_END_COL = 107
S1_FEATURE_ROW_START = 3
S1_FEATURE_ROW_END = 34
S1_PA1_ROW = 36
S1_PAALL_ROW = 37

# Sheet2 layout
S2_WORD_COL = 1
S2_ORTHO_COL = 2
S2_AVGTIME_COL = 3
S2_CODES_START_COL = 4
S2_CODES_END_COL = 109
S2_ROW_START = 1
S2_ROW_END = 60

# Load workbook
wb = load_workbook(FILE_PATH, data_only=True)
ws1 = wb[SHEET1]
ws2 = wb[SHEET2]

# ----------------------------
# Read Sheet1
# ----------------------------
sheet1_rows = []
for r in range(S1_FEATURE_ROW_START, S1_FEATURE_ROW_END + 1):
    vals = []
    for c in range(S1_START_COL, S1_END_COL + 1):
        vals.append(ws1.cell(row=r, column=c).value)
    sheet1_rows.append(vals)

sheet1_rows = np.array(sheet1_rows, dtype=float)
X_sheet1 = sheet1_rows.T   # (106, 32)

y_pa1 = []
y_paall = []
for c in range(S1_START_COL, S1_END_COL + 1):
    y_pa1.append(ws1.cell(row=S1_PA1_ROW, column=c).value)
    y_paall.append(ws1.cell(row=S1_PAALL_ROW, column=c).value)

y_pa1 = np.array(y_pa1, dtype=float)
y_paall = np.array(y_paall, dtype=float)

T = X_sheet1[:, :30]
age = X_sheet1[:, 30]
habit = X_sheet1[:, 31]

# ----------------------------
# Read Sheet2
# ----------------------------
words = []
ortho = []
avg_time = []
code_rows = []

for r in range(S2_ROW_START, S2_ROW_END + 1):
    word = ws2.cell(row=r, column=S2_WORD_COL).value
    osyll = ws2.cell(row=r, column=S2_ORTHO_COL).value
    tsec = ws2.cell(row=r, column=S2_AVGTIME_COL).value

    if word is not None and osyll is not None and tsec is not None:
        words.append(word)
        ortho.append(float(osyll))
        avg_time.append(float(tsec))

        row_codes = []
        for c in range(S2_CODES_START_COL, S2_CODES_END_COL + 1):
            row_codes.append(ws2.cell(row=r, column=c).value)
        code_rows.append(row_codes)

words = words[:30]
ortho = np.array(ortho[:30], dtype=float)
avg_time = np.array(avg_time[:30], dtype=float)
code_rows = np.array(code_rows[:30], dtype=float)

codes_pw = code_rows.T   # (106, 30)

print("Loaded successfully from:", FILE_PATH)
print("T shape:", T.shape)
print("age shape:", age.shape)
print("habit shape:", habit.shape)
print("ortho shape:", ortho.shape)
print("avg_time shape:", avg_time.shape)
print("codes_pw shape:", codes_pw.shape)
print("y_pa1 shape:", y_pa1.shape)
print("y_paall shape:", y_paall.shape)





# ============================================================
# IMPORTS
# ============================================================
import numpy as np
import pandas as pd

from sklearn.model_selection import train_test_split, KFold, GridSearchCV
from sklearn.preprocessing import StandardScaler
from sklearn.pipeline import Pipeline
from sklearn.metrics import r2_score, mean_absolute_error, mean_squared_error

from sklearn.impute import SimpleImputer
from sklearn.linear_model import LinearRegression, Ridge, Lasso, ElasticNet
from sklearn.svm import SVR
from sklearn.ensemble import (
    GradientBoostingRegressor,
    RandomForestRegressor,
    AdaBoostRegressor,
    VotingRegressor
)
from sklearn.neighbors import KNeighborsRegressor
from sklearn.neural_network import MLPRegressor

from lightgbm import LGBMRegressor
import xgboost as xgb
# ============================================================
# TRAIN TEST SPLIT
# ============================================================
train_idx, test_idx = train_test_split(
    np.arange(106),
    test_size=20,
    random_state=42
)
# ============================================================
# FEATURE ENGINEERING (ALL FEATURES INCLUDED)
# ============================================================
eps = 1e-8

def build_all_features(T, age, habit, ortho, avg_time, codes):

    n = T.shape[0]

    mean_t = np.mean(T, axis=1)
    std_t = np.std(T, axis=1)
    median_t = np.median(T, axis=1)
    max_t = np.max(T, axis=1)
    min_t = np.min(T, axis=1)

    # normalized
    norm = T / (avg_time.reshape(1,-1) + eps)

    # residual
    resid = T - avg_time.reshape(1,-1)

    # attempts
    attempts = np.where(codes<=10, codes, codes-10)

    # ==== STRONG FEATURES ====
    slow_ratio = np.mean(T > median_t[:,None], axis=1)
    extreme_ratio = np.mean(T > (mean_t[:,None] + 2*std_t[:,None]), axis=1)
    percentile_gap = np.percentile(T,90,axis=1) - np.percentile(T,10,axis=1)

    corr_ortho = np.array([
        np.corrcoef(T[i], ortho)[0,1] if np.std(T[i])>0 else 0
        for i in range(n)
    ])

    slope = np.array([
        np.polyfit(np.arange(30), T[i], 1)[0]
        for i in range(n)
    ])

    # ==== ATTEMPT FEATURES ====
    mean_attempt = np.mean(attempts, axis=1)
    max_attempt = np.max(attempts, axis=1)
    fail_rate = np.mean(codes>=11, axis=1)

    # ==== COMBINE ====
    X = pd.DataFrame({
        "mean_t": mean_t,
        "std_t": std_t,
        "median_t": median_t,
        "max_t": max_t,
        "min_t": min_t,

        "slow_ratio": slow_ratio,
        "extreme_ratio": extreme_ratio,
        "percentile_gap": percentile_gap,

        "mean_norm": np.mean(norm, axis=1),
        "std_norm": np.std(norm, axis=1),

        "mean_resid": np.mean(resid, axis=1),
        "std_resid": np.std(resid, axis=1),

        "corr_ortho": corr_ortho,
        "slope": slope,

        "mean_attempt": mean_attempt,
        "max_attempt": max_attempt,
        "fail_rate": fail_rate,

        "age": age,
        "habit": habit,

        "age_time": age * mean_t,
        "habit_time": habit * mean_t,
        "efficiency": mean_t / (habit + eps)
    })

    return X.fillna(0)

# ============================================================
# BUILD FEATURES
# ============================================================
X_train = build_all_features(
    T[train_idx], age[train_idx], habit[train_idx],
    ortho, avg_time, codes_pw[train_idx]
)

X_test = build_all_features(
    T[test_idx], age[test_idx], habit[test_idx],
    ortho, avg_time, codes_pw[test_idx]
)

# ============================================================
# TARGET (LOG TRANSFORM)
# ============================================================
y_train = np.log1p(y_pa1[train_idx])
y_test  = np.log1p(y_pa1[test_idx])


# ============================================================
# MODELS
# ============================================================
def get_models():
    return {
        "LinearRegression": (
            Pipeline([
                ("imputer", SimpleImputer(strategy="median")),
                ("scaler", StandardScaler()),
                ("model", LinearRegression())
            ]),
            {}
        ),

        "Ridge": (
            Pipeline([
                ("imputer", SimpleImputer(strategy="median")),
                ("scaler", StandardScaler()),
                ("model", Ridge())
            ]),
            {
                "model__alpha": [0.1, 1, 5, 10, 20, 50, 100]
            }
        ),

        "Lasso": (
            Pipeline([
                ("imputer", SimpleImputer(strategy="median")),
                ("scaler", StandardScaler()),
                ("model", Lasso(max_iter=5000))
            ]),
            {
                "model__alpha": [0.0001, 0.001, 0.01, 0.1, 1]
            }
        ),

        "SVR": (
            Pipeline([
                ("imputer", SimpleImputer(strategy="median")),
                ("scaler", StandardScaler()),
                ("model", SVR())
            ]),
            {
                "model__kernel": ["rbf", "linear"],
                "model__C": [0.1, 1, 5, 10],
                "model__epsilon": [0.01, 0.05, 0.1]
            }
        ),

        "KNN": (
            Pipeline([
                ("imputer", SimpleImputer(strategy="median")),
                ("scaler", StandardScaler()),
                ("model", KNeighborsRegressor())
            ]),
            {
                "model__n_neighbors": [3, 5, 7, 9],
                "model__weights": ["uniform", "distance"]
            }
        ),

        "RandomForest": (
            Pipeline([
                ("imputer", SimpleImputer(strategy="median")),
                ("model", RandomForestRegressor(random_state=42))
            ]),
            {
                "model__n_estimators": [100, 200],
                "model__max_depth": [None, 3, 5, 10]
            }
        ),

        "GradientBoosting": (
            Pipeline([
                ("imputer", SimpleImputer(strategy="median")),
                ("model", GradientBoostingRegressor(random_state=42))
            ]),
            {
                "model__n_estimators": [100, 150, 200],
                "model__learning_rate": [0.03, 0.05, 0.1],
                "model__max_depth": [2, 3]
            }
        ),

        "AdaBoost": (
            Pipeline([
                ("imputer", SimpleImputer(strategy="median")),
                ("model", AdaBoostRegressor(random_state=42))
            ]),
            {
                "model__n_estimators": [50, 100, 200],
                "model__learning_rate": [0.01, 0.05, 0.1, 0.5]
            }
        ),

        "ElasticNet": (
            Pipeline([
                ("imputer", SimpleImputer(strategy="median")),
                ("scaler", StandardScaler()),
                ("model", ElasticNet(max_iter=5000))
            ]),
            {
                "model__alpha": [0.0001, 0.001, 0.01, 0.1, 1],
                "model__l1_ratio": [0.1, 0.3, 0.5, 0.7, 0.9]
            }
        ),

        "LGBM": (
            Pipeline([
                ("imputer", SimpleImputer(strategy="median")),
                ("model", LGBMRegressor(random_state=42))
            ]),
            {
                "model__n_estimators": [100, 200],
                "model__learning_rate": [0.03, 0.05, 0.1],
                "model__num_leaves": [15, 31, 63]
            }
        ),

        "XGBoost": (
            Pipeline([
                ("model", xgb.XGBRegressor(random_state=42, objective="reg:squarederror", verbosity=0))
            ]),
            {
                "model__n_estimators": [50, 100, 200],
                "model__learning_rate": [0.01, 0.03, 0.05, 0.1],
                "model__max_depth": [2, 3, 5],
                "model__subsample": [0.8, 1.0],
                "model__colsample_bytree": [0.8, 1.0]
            }
        ),

        VotingEnsemble": (
            Pipeline([
                ("model", VotingRegressor([
                    ("gb", GradientBoostingRegressor(random_state=42, n_estimators=100, learning_rate=0.05, max_depth=3)),
                    ("rf", RandomForestRegressor(random_state=42, n_estimators=200)),
                    ("xgb", xgb.XGBRegressor(random_state=42, objective="reg:squarederror", verbosity=0, n_estimators=100))
                ]))
            ]),
            {}
        )
    }

# ============================================================
# TRAIN + EVALUATE
# ============================================================
def run_models(X_train, X_test, y_train, y_test):
    cv = KFold(n_splits=5, shuffle=True, random_state=42)
    models = get_models()

    rows = []
    preds = {}

    for name, (pipe, grid) in models.items():
        gs = GridSearchCV(
            estimator=pipe,
            param_grid=grid,
            scoring="r2",
            cv=cv,
            n_jobs=-1,
            refit=True
        )

        gs.fit(X_train, y_train)
        pred_log = gs.best_estimator_.predict(X_test)

        pred = np.expm1(pred_log)
        true = np.expm1(y_test)

        r2 = r2_score(true, pred)
        mae = mean_absolute_error(true, pred)
        rmse = np.sqrt(mean_squared_error(true, pred))

        preds[name] = pred

        rows.append({
            "Model": name,
            "BestCV_R2_log": gs.best_score_,
            "Test_R2": r2,
            "Test_MAE": mae,
            "Test_RMSE": rmse,
            "BestParams": str(gs.best_params_)
        })

        print(f"{name:18s} | CV R2(log)={gs.best_score_:.4f} | Test R2={r2:.4f} | RMSE={rmse:.4f}")

    results_df = pd.DataFrame(rows).sort_values(by="Test_R2", ascending=False).reset_index(drop=True)

    # simple ensemble over top 3 models by test R2
    top3 = results_df["Model"].iloc[:3].tolist()
    ensemble_pred = np.mean([preds[m] for m in top3], axis=0)
    ensemble_r2 = r2_score(np.expm1(y_test), ensemble_pred)
    ensemble_mae = mean_absolute_error(np.expm1(y_test), ensemble_pred)
    ensemble_rmse = np.sqrt(mean_squared_error(np.expm1(y_test), ensemble_pred))

    print("\nTop 3 ensemble:", top3)
    print(f"Ensemble Test R2={ensemble_r2:.4f} | MAE={ensemble_mae:.4f} | RMSE={ensemble_rmse:.4f}")

    return results_df

results_df = run_models(X_train, X_test, y_train, y_test)
print(results_df.head(12))
