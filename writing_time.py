# ============================================================
# LINEAR-REGRESSION-ONLY WRITING TIME PREDICTION
# ============================================================

import warnings
warnings.filterwarnings("ignore")

import numpy as np
import pandas as pd
from openpyxl import load_workbook

from sklearn.model_selection import train_test_split
from sklearn.impute import SimpleImputer
from sklearn.pipeline import Pipeline
from sklearn.preprocessing import StandardScaler
from sklearn.metrics import mean_absolute_error, mean_squared_error, r2_score
from sklearn.linear_model import LinearRegression

# ============================================================
# 1. FILE SETTINGS
# ============================================================
FILE_PATH = "<insert file path here>"
SHEET1 = "Sheet1"
SHEET2 = "Sheet2"

# ============================================================
# 2. EXCEL LAYOUT
# ============================================================
S1_START_COL = 2
S1_END_COL = 107
S1_FEATURE_ROW_START = 3
S1_FEATURE_ROW_END = 34

S2_WORD_COL = 1
S2_ORTHO_COL = 2
S2_AVGTIME_COL = 3
S2_CODES_START_COL = 4
S2_CODES_END_COL = 109
S2_ROW_START = 1
S2_ROW_END = 60

# ============================================================
# 3. LOAD WORKBOOK
# ============================================================
wb = load_workbook(FILE_PATH, data_only=True)
ws1 = wb[SHEET1]
ws2 = wb[SHEET2]

# ============================================================
# 4. READ SHEET1
# ============================================================
sheet1_rows = []
for r in range(S1_FEATURE_ROW_START, S1_FEATURE_ROW_END + 1):
    vals = []
    for c in range(S1_START_COL, S1_END_COL + 1):
        vals.append(ws1.cell(row=r, column=c).value)
    sheet1_rows.append(vals)

sheet1_rows = np.array(sheet1_rows, dtype=float)   # (32, 106)
X_sheet1 = sheet1_rows.T                           # (106, 32)

T_raw = X_sheet1[:, :30]
age = X_sheet1[:, 30]
habit = X_sheet1[:, 31]

print("Sheet1 loaded")
print("X_sheet1 shape:", X_sheet1.shape)
print("T_raw shape:", T_raw.shape)
print("age shape:", age.shape)
print("habit shape:", habit.shape)

# ============================================================
# 5. READ SHEET2
# ============================================================
words = []
ortho = []
avg_word_time = []
code_rows = []

for r in range(S2_ROW_START, S2_ROW_END + 1):
    word = ws2.cell(row=r, column=S2_WORD_COL).value
    osyll = ws2.cell(row=r, column=S2_ORTHO_COL).value
    tsec = ws2.cell(row=r, column=S2_AVGTIME_COL).value

    if word is not None and osyll is not None and tsec is not None:
        words.append(str(word))
        ortho.append(float(osyll))
        avg_word_time.append(float(tsec))

        row_codes = []
        for c in range(S2_CODES_START_COL, S2_CODES_END_COL + 1):
            row_codes.append(ws2.cell(row=r, column=c).value)
        code_rows.append(row_codes)

words = words[:30]
ortho = np.array(ortho[:30], dtype=float)
avg_word_time = np.array(avg_word_time[:30], dtype=float)
code_rows = np.array(code_rows[:30], dtype=float)

assert len(words) == 30, f"Expected 30 words, got {len(words)}"
assert code_rows.shape == (30, 106), f"Expected (30,106), got {code_rows.shape}"

codes_pw = code_rows.T  # (106, 30)

print("\nSheet2 loaded")
print("Number of words:", len(words))
print("ortho shape:", ortho.shape)
print("avg_word_time shape:", avg_word_time.shape)
print("codes_pw shape:", codes_pw.shape)

# ============================================================
# 6. HELPERS
# ============================================================
eps = 1e-8

def code_to_attempt(code):
    try:
        code = int(code)
    except:
        return np.nan

    if code in [1, 11]:
        return 1
    elif code in [2, 12]:
        return 2
    elif code in [3, 13]:
        return 3
    elif code in [4, 14]:
        return 4
    return np.nan

def safe_corr(a, b):
    a = np.asarray(a, dtype=float)
    b = np.asarray(b, dtype=float)
    if np.std(a) < 1e-12 or np.std(b) < 1e-12:
        return 0.0
    return np.corrcoef(a, b)[0, 1]

def linear_slope(x, y):
    x = np.asarray(x, dtype=float)
    y = np.asarray(y, dtype=float)
    if np.std(y) < 1e-12:
        return 0.0
    return np.polyfit(x, y, 1)[0]

def compute_train_amnesia_features(codes_pw_subset):
    n_participants, n_words = codes_pw_subset.shape

    am1_train = np.zeros(n_participants, dtype=float)
    amall_train = np.zeros(n_participants, dtype=float)
    mean_attempt_train = np.zeros(n_participants, dtype=float)
    max_attempt_train = np.zeros(n_participants, dtype=float)
    fail_rate_train = np.zeros(n_participants, dtype=float)
    delayed_rate_train = np.zeros(n_participants, dtype=float)

    for p in range(n_participants):
        row = codes_pw_subset[p]

        count_first_success = np.sum(row == 1)
        am1_train[p] = (n_words - count_first_success) / n_words

        count_total_fail = np.sum(np.isin(row, [11, 12, 13, 14]))
        amall_train[p] = count_total_fail / n_words

        attempts = np.array([code_to_attempt(x) for x in row], dtype=float)
        if np.all(np.isnan(attempts)):
            mean_attempt_train[p] = 0.0
            max_attempt_train[p] = 0.0
        else:
            mean_attempt_train[p] = np.nanmean(attempts)
            max_attempt_train[p] = np.nanmax(attempts)

        fail_rate_train[p] = count_total_fail / n_words
        delayed_rate_train[p] = np.sum(np.isin(row, [2, 3, 4])) / n_words

    return (
        am1_train,
        amall_train,
        mean_attempt_train,
        max_attempt_train,
        fail_rate_train,
        delayed_rate_train
    )

# ============================================================
# 7. WORD SPLIT
# ============================================================
rng = np.random.RandomState(42)
word_idx = np.arange(30)
rng.shuffle(word_idx)

train_word_idx = word_idx[:24]
test_word_idx = word_idx[24:30]

print("\nWord split")
print("Train words:", len(train_word_idx))
print("Test words:", len(test_word_idx))

# ============================================================
# 8. TRAIN-WORD ATTEMPT FEATURES
# ============================================================
codes_train = codes_pw[:, train_word_idx]

(
    am1_train,
    amall_train,
    mean_attempt_train,
    max_attempt_train,
    fail_rate_train,
    delayed_rate_train
) = compute_train_amnesia_features(codes_train)

# ============================================================
# 9. TARGET
# ============================================================
# Raw proxy target
y_pt_proxy = np.mean(T_raw[:, test_word_idx], axis=1)

# ============================================================
# 10. PARTICIPANT SPLIT FIRST
# ============================================================
participant_ids = np.arange(1, 107)

train_idx, test_idx = train_test_split(
    np.arange(106),
    test_size=20,
    random_state=42
)

# training-only reference average to reduce leakage
avg_word_time_train_ref = np.mean(T_raw[train_idx], axis=0)

# ============================================================
# 11. FEATURE ENGINEERING
# ============================================================
def build_participant_features(
    T_raw, age, habit,
    am1_train, amall_train,
    mean_attempt_train, max_attempt_train, fail_rate_train, delayed_rate_train,
    ortho, avg_word_time_ref, train_word_idx, codes_pw
):
    Ttr = T_raw[:, train_word_idx]
    ortho_tr = ortho[train_word_idx]
    avg_time_tr = avg_word_time_ref[train_word_idx]

    n = Ttr.shape[0]
    idx = np.arange(1, Ttr.shape[1] + 1)

    attempts_tr = np.array([
        [code_to_attempt(x) for x in row]
        for row in codes_pw[:, train_word_idx]
    ], dtype=float)
    attempts_tr = np.nan_to_num(attempts_tr, nan=0.0)

    mean_t = np.mean(Ttr, axis=1)
    std_t = np.std(Ttr, axis=1)
    median_t = np.median(Ttr, axis=1)
    min_t = np.min(Ttr, axis=1)
    max_t = np.max(Ttr, axis=1)
    range_t = max_t - min_t
    cv_t = std_t / (mean_t + eps)

    slow_ratio = np.mean(Ttr > median_t[:, None], axis=1)
    extreme_ratio = np.mean(Ttr > (mean_t[:, None] + 2 * std_t[:, None]), axis=1)
    percentile_gap = np.percentile(Ttr, 90, axis=1) - np.percentile(Ttr, 10, axis=1)

    first8 = np.mean(Ttr[:, :8], axis=1)
    middle8 = np.mean(Ttr[:, 8:16], axis=1)
    last8 = np.mean(Ttr[:, 16:24], axis=1)
    late_minus_early = last8 - first8
    slope_trial = np.array([linear_slope(idx, Ttr[i]) for i in range(n)])

    norm_t = Ttr / (avg_time_tr.reshape(1, -1) + eps)
    mean_norm = np.mean(norm_t, axis=1)
    std_norm = np.std(norm_t, axis=1)
    max_norm = np.max(norm_t, axis=1)

    resid_t = Ttr - avg_time_tr.reshape(1, -1)
    mean_resid = np.mean(resid_t, axis=1)
    std_resid = np.std(resid_t, axis=1)

    weighted_ortho_time = np.sum(Ttr * ortho_tr.reshape(1, -1), axis=1) / (np.sum(ortho_tr) + eps)
    corr_time_ortho = np.array([safe_corr(Ttr[i], ortho_tr) for i in range(n)])
    corr_time_avg = np.array([safe_corr(Ttr[i], avg_time_tr) for i in range(n)])
    slope_ortho = np.array([linear_slope(ortho_tr, Ttr[i]) for i in range(n)])

    q1 = np.quantile(ortho_tr, 1/3)
    q2 = np.quantile(ortho_tr, 2/3)

    low_mask = ortho_tr <= q1
    mid_mask = (ortho_tr > q1) & (ortho_tr <= q2)
    high_mask = ortho_tr > q2

    mean_low = np.mean(Ttr[:, low_mask], axis=1)
    mean_mid = np.mean(Ttr[:, mid_mask], axis=1)
    mean_high = np.mean(Ttr[:, high_mask], axis=1)
    high_minus_low = mean_high - mean_low
    high_over_low = mean_high / (mean_low + eps)

    time_per_habit = mean_t / (habit + eps)
    age_time_interaction = age * mean_t
    habit_time_interaction = habit * mean_t

    df_feat = pd.DataFrame({
        "age": age,
        "habit": habit,

        "am1_train": am1_train,
        "amall_train": amall_train,
        "mean_attempt_train": mean_attempt_train,
        "max_attempt_train": max_attempt_train,
        "fail_rate_train": fail_rate_train,
        "delayed_rate_train": delayed_rate_train,

        "mean_t": mean_t,
        "std_t": std_t,
        "median_t": median_t,
        "min_t": min_t,
        "max_t": max_t,
        "range_t": range_t,
        "cv_t": cv_t,

        "slow_ratio": slow_ratio,
        "extreme_ratio": extreme_ratio,
        "percentile_gap": percentile_gap,

        "first8_t": first8,
        "middle8_t": middle8,
        "last8_t": last8,
        "late_minus_early": late_minus_early,
        "slope_trial": slope_trial,

        "mean_norm": mean_norm,
        "std_norm": std_norm,
        "max_norm": max_norm,
        "mean_resid": mean_resid,
        "std_resid": std_resid,

        "weighted_ortho_time": weighted_ortho_time,
        "corr_time_ortho": corr_time_ortho,
        "corr_time_avg": corr_time_avg,
        "slope_ortho": slope_ortho,

        "mean_low_ortho_time": mean_low,
        "mean_mid_ortho_time": mean_mid,
        "mean_high_ortho_time": mean_high,
        "high_minus_low": high_minus_low,
        "high_over_low": high_over_low,

        "time_per_habit": time_per_habit,
        "age_time_interaction": age_time_interaction,
        "habit_time_interaction": habit_time_interaction
    })

    # per-word raw times
    for j in range(Ttr.shape[1]):
        df_feat[f"raw_time_{j+1:02d}"] = Ttr[:, j]

    # per-word normalized times
    for j in range(norm_t.shape[1]):
        df_feat[f"norm_time_{j+1:02d}"] = norm_t[:, j]

    # per-word attempts
    for j in range(attempts_tr.shape[1]):
        df_feat[f"attempt_{j+1:02d}"] = attempts_tr[:, j]

    return df_feat.fillna(0)

X_feat = build_participant_features(
    T_raw, age, habit,
    am1_train, amall_train,
    mean_attempt_train, max_attempt_train, fail_rate_train, delayed_rate_train,
    ortho, avg_word_time_train_ref, train_word_idx, codes_pw
)

# ============================================================
# 12. TRAIN / TEST MATRICES
# ============================================================
X_train_full = X_feat.iloc[train_idx].reset_index(drop=True)
X_test_full = X_feat.iloc[test_idx].reset_index(drop=True)

y_train = y_pt_proxy[train_idx]
y_test = y_pt_proxy[test_idx]

id_test = participant_ids[test_idx]

# ============================================================
# 13. TRAIN-ONLY FEATURE SELECTION
# ============================================================
def select_top_features(X_df, y, top_k=40):
    corrs = []
    for col in X_df.columns:
        c = safe_corr(X_df[col].values, y)
        corrs.append(abs(c))
    corrs = np.array(corrs)
    idx = np.argsort(corrs)[-top_k:]
    cols = list(X_df.columns[idx])
    return cols, pd.Series(corrs, index=X_df.columns).sort_values(ascending=False)

TOP_K = 40
selected_cols, corr_series = select_top_features(X_train_full, y_train, top_k=TOP_K)

X_train = X_train_full[selected_cols].copy()
X_test = X_test_full[selected_cols].copy()

print("\nSelected feature matrix shape (train):", X_train.shape)
print("Selected feature matrix shape (test):", X_test.shape)
print("\nTop correlated training features:")
print(corr_series.head(15))

# ============================================================
# 14. LINEAR REGRESSION ONLY
# ============================================================
linreg_pipe = Pipeline([
    ("imputer", SimpleImputer(strategy="median")),
    ("scaler", StandardScaler()),
    ("model", LinearRegression())
])

print("\n==================== LINEAR REGRESSION ONLY ====================")
linreg_pipe.fit(X_train, y_train)

y_pred = linreg_pipe.predict(X_test)

mae = mean_absolute_error(y_test, y_pred)
mse = mean_squared_error(y_test, y_pred)
rmse = np.sqrt(mse)
r2 = r2_score(y_test, y_pred)

print("\nModel parameters:")
print("fit_intercept=True, copy_X=True, tol=1e-06, n_jobs=None, positive=False")

print("\nMetrics:")
print(f"Test_MAE  = {mae:.4f}")
print(f"Test_MSE  = {mse:.4f}")
print(f"Test_RMSE = {rmse:.4f}")
print(f"Test_R2   = {r2:.4f}")

pred_df = pd.DataFrame({
    "participant_id": id_test,
    "true_avg_time_test6": y_test,
    "pred_avg_time_test6": y_pred
})

print("\nPredictions preview:")
print(pred_df.head())

# Optional save
# pred_df.to_csv("/content/linear_regression_only_predictions.csv", index=False)
# print("\nSaved: /content/linear_regression_only_predictions.csv")

# ============================================================
# RIDGE-ONLY WRITING TIME PREDICTION
# ============================================================

import warnings
warnings.filterwarnings("ignore")

import numpy as np
import pandas as pd
from openpyxl import load_workbook

from sklearn.model_selection import train_test_split, KFold, GridSearchCV
from sklearn.impute import SimpleImputer
from sklearn.pipeline import Pipeline
from sklearn.preprocessing import StandardScaler
from sklearn.metrics import mean_absolute_error, mean_squared_error, r2_score
from sklearn.linear_model import Ridge

# ============================================================
# 1. FILE SETTINGS
# ============================================================
FILE_PATH = "<insert file path here>"
SHEET1 = "Sheet1"
SHEET2 = "Sheet2"

# ============================================================
# 2. EXCEL LAYOUT
# ============================================================
S1_START_COL = 2
S1_END_COL = 107
S1_FEATURE_ROW_START = 3
S1_FEATURE_ROW_END = 34

S2_WORD_COL = 1
S2_ORTHO_COL = 2
S2_AVGTIME_COL = 3
S2_CODES_START_COL = 4
S2_CODES_END_COL = 109
S2_ROW_START = 1
S2_ROW_END = 60

# ============================================================
# 3. LOAD WORKBOOK
# ============================================================
wb = load_workbook(FILE_PATH, data_only=True)
ws1 = wb[SHEET1]
ws2 = wb[SHEET2]

# ============================================================
# 4. READ SHEET1
# ============================================================
sheet1_rows = []
for r in range(S1_FEATURE_ROW_START, S1_FEATURE_ROW_END + 1):
    vals = []
    for c in range(S1_START_COL, S1_END_COL + 1):
        vals.append(ws1.cell(row=r, column=c).value)
    sheet1_rows.append(vals)

sheet1_rows = np.array(sheet1_rows, dtype=float)   # (32, 106)
X_sheet1 = sheet1_rows.T                           # (106, 32)

T_raw = X_sheet1[:, :30]
age = X_sheet1[:, 30]
habit = X_sheet1[:, 31]

print("Sheet1 loaded")
print("X_sheet1 shape:", X_sheet1.shape)
print("T_raw shape:", T_raw.shape)
print("age shape:", age.shape)
print("habit shape:", habit.shape)

# ============================================================
# 5. READ SHEET2
# ============================================================
words = []
ortho = []
avg_word_time = []
code_rows = []

for r in range(S2_ROW_START, S2_ROW_END + 1):
    word = ws2.cell(row=r, column=S2_WORD_COL).value
    osyll = ws2.cell(row=r, column=S2_ORTHO_COL).value
    tsec = ws2.cell(row=r, column=S2_AVGTIME_COL).value

    if word is not None and osyll is not None and tsec is not None:
        words.append(str(word))
        ortho.append(float(osyll))
        avg_word_time.append(float(tsec))

        row_codes = []
        for c in range(S2_CODES_START_COL, S2_CODES_END_COL + 1):
            row_codes.append(ws2.cell(row=r, column=c).value)
        code_rows.append(row_codes)

words = words[:30]
ortho = np.array(ortho[:30], dtype=float)
avg_word_time = np.array(avg_word_time[:30], dtype=float)
code_rows = np.array(code_rows[:30], dtype=float)

assert len(words) == 30, f"Expected 30 words, got {len(words)}"
assert code_rows.shape == (30, 106), f"Expected (30,106), got {code_rows.shape}"

codes_pw = code_rows.T  # (106, 30)

print("\nSheet2 loaded")
print("Number of words:", len(words))
print("ortho shape:", ortho.shape)
print("avg_word_time shape:", avg_word_time.shape)
print("codes_pw shape:", codes_pw.shape)

# ============================================================
# 6. HELPERS
# ============================================================
eps = 1e-8

def code_to_attempt(code):
    try:
        code = int(code)
    except:
        return np.nan

    if code in [1, 11]:
        return 1
    elif code in [2, 12]:
        return 2
    elif code in [3, 13]:
        return 3
    elif code in [4, 14]:
        return 4
    return np.nan

def safe_corr(a, b):
    a = np.asarray(a, dtype=float)
    b = np.asarray(b, dtype=float)
    if np.std(a) < 1e-12 or np.std(b) < 1e-12:
        return 0.0
    return np.corrcoef(a, b)[0, 1]

def linear_slope(x, y):
    x = np.asarray(x, dtype=float)
    y = np.asarray(y, dtype=float)
    if np.std(y) < 1e-12:
        return 0.0
    return np.polyfit(x, y, 1)[0]

def compute_train_amnesia_features(codes_pw_subset):
    n_participants, n_words = codes_pw_subset.shape

    am1_train = np.zeros(n_participants, dtype=float)
    amall_train = np.zeros(n_participants, dtype=float)
    mean_attempt_train = np.zeros(n_participants, dtype=float)
    max_attempt_train = np.zeros(n_participants, dtype=float)
    fail_rate_train = np.zeros(n_participants, dtype=float)
    delayed_rate_train = np.zeros(n_participants, dtype=float)

    for p in range(n_participants):
        row = codes_pw_subset[p]

        count_first_success = np.sum(row == 1)
        am1_train[p] = (n_words - count_first_success) / n_words

        count_total_fail = np.sum(np.isin(row, [11, 12, 13, 14]))
        amall_train[p] = count_total_fail / n_words

        attempts = np.array([code_to_attempt(x) for x in row], dtype=float)
        if np.all(np.isnan(attempts)):
            mean_attempt_train[p] = 0.0
            max_attempt_train[p] = 0.0
        else:
            mean_attempt_train[p] = np.nanmean(attempts)
            max_attempt_train[p] = np.nanmax(attempts)

        fail_rate_train[p] = count_total_fail / n_words
        delayed_rate_train[p] = np.sum(np.isin(row, [2, 3, 4])) / n_words

    return (
        am1_train,
        amall_train,
        mean_attempt_train,
        max_attempt_train,
        fail_rate_train,
        delayed_rate_train
    )

# ============================================================
# 7. WORD SPLIT
# ============================================================
rng = np.random.RandomState(42)
word_idx = np.arange(30)
rng.shuffle(word_idx)

train_word_idx = word_idx[:24]
test_word_idx = word_idx[24:30]

print("\nWord split")
print("Train words:", len(train_word_idx))
print("Test words:", len(test_word_idx))

# ============================================================
# 8. TRAIN-WORD ATTEMPT FEATURES
# ============================================================
codes_train = codes_pw[:, train_word_idx]

(
    am1_train,
    amall_train,
    mean_attempt_train,
    max_attempt_train,
    fail_rate_train,
    delayed_rate_train
) = compute_train_amnesia_features(codes_train)

# ============================================================
# 9. TARGET
# ============================================================
y_pt_proxy = np.mean(T_raw[:, test_word_idx], axis=1)

# ============================================================
# 10. PARTICIPANT SPLIT FIRST
# ============================================================
participant_ids = np.arange(1, 107)

train_idx, test_idx = train_test_split(
    np.arange(106),
    test_size=20,
    random_state=42
)

# training-only reference average to reduce leakage
avg_word_time_train_ref = np.mean(T_raw[train_idx], axis=0)

# ============================================================
# 11. FEATURE ENGINEERING
# ============================================================
def build_participant_features(
    T_raw, age, habit,
    am1_train, amall_train,
    mean_attempt_train, max_attempt_train, fail_rate_train, delayed_rate_train,
    ortho, avg_word_time_ref, train_word_idx, codes_pw
):
    Ttr = T_raw[:, train_word_idx]
    ortho_tr = ortho[train_word_idx]
    avg_time_tr = avg_word_time_ref[train_word_idx]

    n = Ttr.shape[0]
    idx = np.arange(1, Ttr.shape[1] + 1)

    attempts_tr = np.array([
        [code_to_attempt(x) for x in row]
        for row in codes_pw[:, train_word_idx]
    ], dtype=float)
    attempts_tr = np.nan_to_num(attempts_tr, nan=0.0)

    mean_t = np.mean(Ttr, axis=1)
    std_t = np.std(Ttr, axis=1)
    median_t = np.median(Ttr, axis=1)
    min_t = np.min(Ttr, axis=1)
    max_t = np.max(Ttr, axis=1)
    range_t = max_t - min_t
    cv_t = std_t / (mean_t + eps)

    slow_ratio = np.mean(Ttr > median_t[:, None], axis=1)
    extreme_ratio = np.mean(Ttr > (mean_t[:, None] + 2 * std_t[:, None]), axis=1)
    percentile_gap = np.percentile(Ttr, 90, axis=1) - np.percentile(Ttr, 10, axis=1)

    first8 = np.mean(Ttr[:, :8], axis=1)
    middle8 = np.mean(Ttr[:, 8:16], axis=1)
    last8 = np.mean(Ttr[:, 16:24], axis=1)
    late_minus_early = last8 - first8
    slope_trial = np.array([linear_slope(idx, Ttr[i]) for i in range(n)])

    norm_t = Ttr / (avg_time_tr.reshape(1, -1) + eps)
    mean_norm = np.mean(norm_t, axis=1)
    std_norm = np.std(norm_t, axis=1)
    max_norm = np.max(norm_t, axis=1)

    resid_t = Ttr - avg_time_tr.reshape(1, -1)
    mean_resid = np.mean(resid_t, axis=1)
    std_resid = np.std(resid_t, axis=1)

    weighted_ortho_time = np.sum(Ttr * ortho_tr.reshape(1, -1), axis=1) / (np.sum(ortho_tr) + eps)
    corr_time_ortho = np.array([safe_corr(Ttr[i], ortho_tr) for i in range(n)])
    corr_time_avg = np.array([safe_corr(Ttr[i], avg_time_tr) for i in range(n)])
    slope_ortho = np.array([linear_slope(ortho_tr, Ttr[i]) for i in range(n)])

    q1 = np.quantile(ortho_tr, 1/3)
    q2 = np.quantile(ortho_tr, 2/3)

    low_mask = ortho_tr <= q1
    mid_mask = (ortho_tr > q1) & (ortho_tr <= q2)
    high_mask = ortho_tr > q2

    mean_low = np.mean(Ttr[:, low_mask], axis=1)
    mean_mid = np.mean(Ttr[:, mid_mask], axis=1)
    mean_high = np.mean(Ttr[:, high_mask], axis=1)
    high_minus_low = mean_high - mean_low
    high_over_low = mean_high / (mean_low + eps)

    time_per_habit = mean_t / (habit + eps)
    age_time_interaction = age * mean_t
    habit_time_interaction = habit * mean_t

    df_feat = pd.DataFrame({
        "age": age,
        "habit": habit,

        "am1_train": am1_train,
        "amall_train": amall_train,
        "mean_attempt_train": mean_attempt_train,
        "max_attempt_train": max_attempt_train,
        "fail_rate_train": fail_rate_train,
        "delayed_rate_train": delayed_rate_train,

        "mean_t": mean_t,
        "std_t": std_t,
        "median_t": median_t,
        "min_t": min_t,
        "max_t": max_t,
        "range_t": range_t,
        "cv_t": cv_t,

        "slow_ratio": slow_ratio,
        "extreme_ratio": extreme_ratio,
        "percentile_gap": percentile_gap,

        "first8_t": first8,
        "middle8_t": middle8,
        "last8_t": last8,
        "late_minus_early": late_minus_early,
        "slope_trial": slope_trial,

        "mean_norm": mean_norm,
        "std_norm": std_norm,
        "max_norm": max_norm,
        "mean_resid": mean_resid,
        "std_resid": std_resid,

        "weighted_ortho_time": weighted_ortho_time,
        "corr_time_ortho": corr_time_ortho,
        "corr_time_avg": corr_time_avg,
        "slope_ortho": slope_ortho,

        "mean_low_ortho_time": mean_low,
        "mean_mid_ortho_time": mean_mid,
        "mean_high_ortho_time": mean_high,
        "high_minus_low": high_minus_low,
        "high_over_low": high_over_low,

        "time_per_habit": time_per_habit,
        "age_time_interaction": age_time_interaction,
        "habit_time_interaction": habit_time_interaction
    })

    for j in range(Ttr.shape[1]):
        df_feat[f"raw_time_{j+1:02d}"] = Ttr[:, j]

    for j in range(norm_t.shape[1]):
        df_feat[f"norm_time_{j+1:02d}"] = norm_t[:, j]

    for j in range(attempts_tr.shape[1]):
        df_feat[f"attempt_{j+1:02d}"] = attempts_tr[:, j]

    return df_feat.fillna(0)

X_feat = build_participant_features(
    T_raw, age, habit,
    am1_train, amall_train,
    mean_attempt_train, max_attempt_train, fail_rate_train, delayed_rate_train,
    ortho, avg_word_time_train_ref, train_word_idx, codes_pw
)

# ============================================================
# 12. TRAIN / TEST MATRICES
# ============================================================
X_train_full = X_feat.iloc[train_idx].reset_index(drop=True)
X_test_full = X_feat.iloc[test_idx].reset_index(drop=True)

y_train = y_pt_proxy[train_idx]
y_test = y_pt_proxy[test_idx]

id_test = participant_ids[test_idx]

# ============================================================
# 13. TRAIN-ONLY FEATURE SELECTION
# ============================================================
def select_top_features(X_df, y, top_k=40):
    corrs = []
    for col in X_df.columns:
        c = safe_corr(X_df[col].values, y)
        corrs.append(abs(c))
    corrs = np.array(corrs)
    idx = np.argsort(corrs)[-top_k:]
    cols = list(X_df.columns[idx])
    return cols, pd.Series(corrs, index=X_df.columns).sort_values(ascending=False)

TOP_K = 40
selected_cols, corr_series = select_top_features(X_train_full, y_train, top_k=TOP_K)

X_train = X_train_full[selected_cols].copy()
X_test = X_test_full[selected_cols].copy()

print("\nSelected feature matrix shape (train):", X_train.shape)
print("Selected feature matrix shape (test):", X_test.shape)
print("\nTop correlated training features:")
print(corr_series.head(15))

# ============================================================
# 14. RIDGE ONLY
# ============================================================
ridge_pipe = Pipeline([
    ("imputer", SimpleImputer(strategy="median")),
    ("scaler", StandardScaler()),
    ("model", Ridge(random_state=42))
])

ridge_param_grid = {}

cv = KFold(n_splits=5, shuffle=True, random_state=42)

search = GridSearchCV(
    estimator=ridge_pipe,
    param_grid=ridge_param_grid,
    scoring="r2",
    cv=cv,
    n_jobs=-1,
    refit=True,
    error_score="raise"
)

print("\n==================== RIDGE ONLY ====================")
search.fit(X_train, y_train)

y_pred = search.best_estimator_.predict(X_test)

mae = mean_absolute_error(y_test, y_pred)
mse = mean_squared_error(y_test, y_pred)
rmse = np.sqrt(mse)
r2 = r2_score(y_test, y_pred)

print("\nBest Params:")
print(search.best_params_)

print("\nMetrics:")
print(f"BestCV_R2 = {search.best_score_:.4f}")
print(f"Test_MAE  = {mae:.4f}")
print(f"Test_MSE  = {mse:.4f}")
print(f"Test_RMSE = {rmse:.4f}")
print(f"Test_R2   = {r2:.4f}")

pred_df = pd.DataFrame({
    "participant_id": id_test,
    "true_avg_time_test6": y_test,
    "pred_avg_time_test6": y_pred
})

print("\nPredictions preview:")
print(pred_df.head())

# Optional save
# pred_df.to_csv("/content/ridge_only_predictions.csv", index=False)
# print("\nSaved: /content/ridge_only_predictions.csv")

# ============================================================
# LASSO-ONLY WRITING TIME PREDICTION
# ============================================================

import warnings
warnings.filterwarnings("ignore")

import numpy as np
import pandas as pd
from openpyxl import load_workbook

from sklearn.model_selection import train_test_split, KFold, GridSearchCV
from sklearn.impute import SimpleImputer
from sklearn.pipeline import Pipeline
from sklearn.preprocessing import StandardScaler
from sklearn.metrics import mean_absolute_error, mean_squared_error, r2_score
from sklearn.linear_model import Lasso

# ============================================================
# 1. FILE SETTINGS
# ============================================================
FILE_PATH ="<insert file path here>"
SHEET1 = "Sheet1"
SHEET2 = "Sheet2"

# ============================================================
# 2. EXCEL LAYOUT
# ============================================================
S1_START_COL = 2
S1_END_COL = 107
S1_FEATURE_ROW_START = 3
S1_FEATURE_ROW_END = 34

S2_WORD_COL = 1
S2_ORTHO_COL = 2
S2_AVGTIME_COL = 3
S2_CODES_START_COL = 4
S2_CODES_END_COL = 109
S2_ROW_START = 1
S2_ROW_END = 60

# ============================================================
# 3. LOAD WORKBOOK
# ============================================================
wb = load_workbook(FILE_PATH, data_only=True)
ws1 = wb[SHEET1]
ws2 = wb[SHEET2]

# ============================================================
# 4. READ SHEET1
# ============================================================
sheet1_rows = []
for r in range(S1_FEATURE_ROW_START, S1_FEATURE_ROW_END + 1):
    vals = []
    for c in range(S1_START_COL, S1_END_COL + 1):
        vals.append(ws1.cell(row=r, column=c).value)
    sheet1_rows.append(vals)

sheet1_rows = np.array(sheet1_rows, dtype=float)   # (32, 106)
X_sheet1 = sheet1_rows.T                           # (106, 32)

T_raw = X_sheet1[:, :30]      # participant x 30 raw time-like features
age = X_sheet1[:, 30]
habit = X_sheet1[:, 31]

print("Sheet1 loaded")
print("X_sheet1 shape:", X_sheet1.shape)
print("T_raw shape:", T_raw.shape)
print("age shape:", age.shape)
print("habit shape:", habit.shape)

# ============================================================
# 5. READ SHEET2
# ============================================================
words = []
ortho = []
avg_word_time = []
code_rows = []

for r in range(S2_ROW_START, S2_ROW_END + 1):
    word = ws2.cell(row=r, column=S2_WORD_COL).value
    osyll = ws2.cell(row=r, column=S2_ORTHO_COL).value
    tsec = ws2.cell(row=r, column=S2_AVGTIME_COL).value

    if word is not None and osyll is not None and tsec is not None:
        words.append(str(word))
        ortho.append(float(osyll))
        avg_word_time.append(float(tsec))

        row_codes = []
        for c in range(S2_CODES_START_COL, S2_CODES_END_COL + 1):
            row_codes.append(ws2.cell(row=r, column=c).value)
        code_rows.append(row_codes)

words = words[:30]
ortho = np.array(ortho[:30], dtype=float)
avg_word_time = np.array(avg_word_time[:30], dtype=float)
code_rows = np.array(code_rows[:30], dtype=float)

assert len(words) == 30, f"Expected 30 words, got {len(words)}"
assert code_rows.shape == (30, 106), f"Expected (30,106), got {code_rows.shape}"

codes_pw = code_rows.T  # (106, 30)

print("\nSheet2 loaded")
print("Number of words:", len(words))
print("ortho shape:", ortho.shape)
print("avg_word_time shape:", avg_word_time.shape)
print("codes_pw shape:", codes_pw.shape)

# ============================================================
# 6. HELPERS
# ============================================================
eps = 1e-8

def code_to_attempt(code):
    try:
        code = int(code)
    except:
        return np.nan

    if code in [1, 11]:
        return 1
    elif code in [2, 12]:
        return 2
    elif code in [3, 13]:
        return 3
    elif code in [4, 14]:
        return 4
    return np.nan

def safe_corr(a, b):
    a = np.asarray(a, dtype=float)
    b = np.asarray(b, dtype=float)
    if np.std(a) < 1e-12 or np.std(b) < 1e-12:
        return 0.0
    return np.corrcoef(a, b)[0, 1]

def linear_slope(x, y):
    x = np.asarray(x, dtype=float)
    y = np.asarray(y, dtype=float)
    if np.std(y) < 1e-12:
        return 0.0
    return np.polyfit(x, y, 1)[0]

def compute_train_amnesia_features(codes_pw_subset):
    n_participants, n_words = codes_pw_subset.shape

    am1_train = np.zeros(n_participants, dtype=float)
    amall_train = np.zeros(n_participants, dtype=float)
    mean_attempt_train = np.zeros(n_participants, dtype=float)
    max_attempt_train = np.zeros(n_participants, dtype=float)
    fail_rate_train = np.zeros(n_participants, dtype=float)
    delayed_rate_train = np.zeros(n_participants, dtype=float)

    for p in range(n_participants):
        row = codes_pw_subset[p]

        count_first_success = np.sum(row == 1)
        am1_train[p] = (n_words - count_first_success) / n_words

        count_total_fail = np.sum(np.isin(row, [11, 12, 13, 14]))
        amall_train[p] = count_total_fail / n_words

        attempts = np.array([code_to_attempt(x) for x in row], dtype=float)
        if np.all(np.isnan(attempts)):
            mean_attempt_train[p] = 0.0
            max_attempt_train[p] = 0.0
        else:
            mean_attempt_train[p] = np.nanmean(attempts)
            max_attempt_train[p] = np.nanmax(attempts)

        fail_rate_train[p] = count_total_fail / n_words
        delayed_rate_train[p] = np.sum(np.isin(row, [2, 3, 4])) / n_words

    return (
        am1_train,
        amall_train,
        mean_attempt_train,
        max_attempt_train,
        fail_rate_train,
        delayed_rate_train
    )

# ============================================================
# 7. WORD SPLIT
# ============================================================
rng = np.random.RandomState(42)
word_idx = np.arange(30)
rng.shuffle(word_idx)

train_word_idx = word_idx[:24]
test_word_idx = word_idx[24:30]

print("\nWord split")
print("Train words:", len(train_word_idx))
print("Test words:", len(test_word_idx))

# ============================================================
# 8. TRAIN-WORD ATTEMPT FEATURES
# ============================================================
codes_train = codes_pw[:, train_word_idx]

(
    am1_train,
    amall_train,
    mean_attempt_train,
    max_attempt_train,
    fail_rate_train,
    delayed_rate_train
) = compute_train_amnesia_features(codes_train)

# ============================================================
# 9. TARGET
# ============================================================
# proxy target = mean raw time-like feature over held-out 6 word positions
y_pt_proxy = np.mean(T_raw[:, test_word_idx], axis=1)

# ============================================================
# 10. PARTICIPANT SPLIT FIRST
# ============================================================
participant_ids = np.arange(1, 107)

train_idx, test_idx = train_test_split(
    np.arange(106),
    test_size=20,
    random_state=42
)

# training-only reference average to reduce leakage
avg_word_time_train_ref = np.mean(T_raw[train_idx], axis=0)

# ============================================================
# 11. FEATURE ENGINEERING
# ============================================================
def build_participant_features(
    T_raw, age, habit,
    am1_train, amall_train,
    mean_attempt_train, max_attempt_train, fail_rate_train, delayed_rate_train,
    ortho, avg_word_time_ref, train_word_idx, codes_pw
):
    Ttr = T_raw[:, train_word_idx]
    ortho_tr = ortho[train_word_idx]
    avg_time_tr = avg_word_time_ref[train_word_idx]

    n = Ttr.shape[0]
    idx = np.arange(1, Ttr.shape[1] + 1)

    attempts_tr = np.array([
        [code_to_attempt(x) for x in row]
        for row in codes_pw[:, train_word_idx]
    ], dtype=float)
    attempts_tr = np.nan_to_num(attempts_tr, nan=0.0)

    mean_t = np.mean(Ttr, axis=1)
    std_t = np.std(Ttr, axis=1)
    median_t = np.median(Ttr, axis=1)
    min_t = np.min(Ttr, axis=1)
    max_t = np.max(Ttr, axis=1)
    range_t = max_t - min_t
    cv_t = std_t / (mean_t + eps)

    slow_ratio = np.mean(Ttr > median_t[:, None], axis=1)
    extreme_ratio = np.mean(Ttr > (mean_t[:, None] + 2 * std_t[:, None]), axis=1)
    percentile_gap = np.percentile(Ttr, 90, axis=1) - np.percentile(Ttr, 10, axis=1)

    first8 = np.mean(Ttr[:, :8], axis=1)
    middle8 = np.mean(Ttr[:, 8:16], axis=1)
    last8 = np.mean(Ttr[:, 16:24], axis=1)
    late_minus_early = last8 - first8
    slope_trial = np.array([linear_slope(idx, Ttr[i]) for i in range(n)])

    norm_t = Ttr / (avg_time_tr.reshape(1, -1) + eps)
    mean_norm = np.mean(norm_t, axis=1)
    std_norm = np.std(norm_t, axis=1)
    max_norm = np.max(norm_t, axis=1)

    resid_t = Ttr - avg_time_tr.reshape(1, -1)
    mean_resid = np.mean(resid_t, axis=1)
    std_resid = np.std(resid_t, axis=1)

    weighted_ortho_time = np.sum(Ttr * ortho_tr.reshape(1, -1), axis=1) / (np.sum(ortho_tr) + eps)
    corr_time_ortho = np.array([safe_corr(Ttr[i], ortho_tr) for i in range(n)])
    corr_time_avg = np.array([safe_corr(Ttr[i], avg_time_tr) for i in range(n)])
    slope_ortho = np.array([linear_slope(ortho_tr, Ttr[i]) for i in range(n)])

    q1 = np.quantile(ortho_tr, 1/3)
    q2 = np.quantile(ortho_tr, 2/3)

    low_mask = ortho_tr <= q1
    mid_mask = (ortho_tr > q1) & (ortho_tr <= q2)
    high_mask = ortho_tr > q2

    mean_low = np.mean(Ttr[:, low_mask], axis=1)
    mean_mid = np.mean(Ttr[:, mid_mask], axis=1)
    mean_high = np.mean(Ttr[:, high_mask], axis=1)
    high_minus_low = mean_high - mean_low
    high_over_low = mean_high / (mean_low + eps)

    time_per_habit = mean_t / (habit + eps)
    age_time_interaction = age * mean_t
    habit_time_interaction = habit * mean_t

    df_feat = pd.DataFrame({
        "age": age,
        "habit": habit,

        "am1_train": am1_train,
        "amall_train": amall_train,
        "mean_attempt_train": mean_attempt_train,
        "max_attempt_train": max_attempt_train,
        "fail_rate_train": fail_rate_train,
        "delayed_rate_train": delayed_rate_train,

        "mean_t": mean_t,
        "std_t": std_t,
        "median_t": median_t,
        "min_t": min_t,
        "max_t": max_t,
        "range_t": range_t,
        "cv_t": cv_t,

        "slow_ratio": slow_ratio,
        "extreme_ratio": extreme_ratio,
        "percentile_gap": percentile_gap,

        "first8_t": first8,
        "middle8_t": middle8,
        "last8_t": last8,
        "late_minus_early": late_minus_early,
        "slope_trial": slope_trial,

        "mean_norm": mean_norm,
        "std_norm": std_norm,
        "max_norm": max_norm,
        "mean_resid": mean_resid,
        "std_resid": std_resid,

        "weighted_ortho_time": weighted_ortho_time,
        "corr_time_ortho": corr_time_ortho,
        "corr_time_avg": corr_time_avg,
        "slope_ortho": slope_ortho,

        "mean_low_ortho_time": mean_low,
        "mean_mid_ortho_time": mean_mid,
        "mean_high_ortho_time": mean_high,
        "high_minus_low": high_minus_low,
        "high_over_low": high_over_low,

        "time_per_habit": time_per_habit,
        "age_time_interaction": age_time_interaction,
        "habit_time_interaction": habit_time_interaction
    })

    # per-word raw times
    for j in range(Ttr.shape[1]):
        df_feat[f"raw_time_{j+1:02d}"] = Ttr[:, j]

    # per-word normalized times
    for j in range(norm_t.shape[1]):
        df_feat[f"norm_time_{j+1:02d}"] = norm_t[:, j]

    # per-word attempts
    for j in range(attempts_tr.shape[1]):
        df_feat[f"attempt_{j+1:02d}"] = attempts_tr[:, j]

    return df_feat.fillna(0)

X_feat = build_participant_features(
    T_raw, age, habit,
    am1_train, amall_train,
    mean_attempt_train, max_attempt_train, fail_rate_train, delayed_rate_train,
    ortho, avg_word_time_train_ref, train_word_idx, codes_pw
)

# ============================================================
# 12. TRAIN / TEST MATRICES
# ============================================================
X_train_full = X_feat.iloc[train_idx].reset_index(drop=True)
X_test_full = X_feat.iloc[test_idx].reset_index(drop=True)

y_train = y_pt_proxy[train_idx]
y_test = y_pt_proxy[test_idx]

id_test = participant_ids[test_idx]

# ============================================================
# 13. TRAIN-ONLY FEATURE SELECTION
# ============================================================
def select_top_features(X_df, y, top_k=40):
    corrs = []
    for col in X_df.columns:
        c = safe_corr(X_df[col].values, y)
        corrs.append(abs(c))
    corrs = np.array(corrs)
    idx = np.argsort(corrs)[-top_k:]
    cols = list(X_df.columns[idx])
    return cols, pd.Series(corrs, index=X_df.columns).sort_values(ascending=False)

TOP_K = 40
selected_cols, corr_series = select_top_features(X_train_full, y_train, top_k=TOP_K)

X_train = X_train_full[selected_cols].copy()
X_test = X_test_full[selected_cols].copy()

print("\nSelected feature matrix shape (train):", X_train.shape)
print("Selected feature matrix shape (test):", X_test.shape)
print("\nTop correlated training features:")
print(corr_series.head(15))

# ============================================================
# 14. LASSO ONLY
# ============================================================
lasso_pipe = Pipeline([
    ("imputer", SimpleImputer(strategy="median")),
    ("scaler", StandardScaler()),
    ("model", Lasso(max_iter=5000))
])

lasso_param_grid = {
    "model__alpha": [0.0001, 0.0003, 0.001, 0.003, 0.01, 0.03, 0.1, 0.3, 1.0]
}

cv = KFold(n_splits=5, shuffle=True, random_state=42)

search = GridSearchCV(
    estimator=lasso_pipe,
    param_grid=lasso_param_grid,
    scoring="r2",
    cv=cv,
    n_jobs=-1,
    refit=True,
    error_score="raise"
)

print("\n==================== LASSO ONLY ====================")
search.fit(X_train, y_train)

y_pred = search.best_estimator_.predict(X_test)

mae = mean_absolute_error(y_test, y_pred)
mse = mean_squared_error(y_test, y_pred)
rmse = np.sqrt(mse)
r2 = r2_score(y_test, y_pred)

print("\nBest Params:")
print(search.best_params_)

print("\nMetrics:")
print(f"BestCV_R2 = {search.best_score_:.4f}")
print(f"Test_MAE  = {mae:.4f}")
print(f"Test_MSE  = {mse:.4f}")
print(f"Test_RMSE = {rmse:.4f}")
print(f"Test_R2   = {r2:.4f}")

pred_df = pd.DataFrame({
    "participant_id": id_test,
    "true_avg_time_test6": y_test,
    "pred_avg_time_test6": y_pred
})

print("\nPredictions preview:")
print(pred_df.head())

# Optional save
# pred_df.to_csv("/content/lasso_only_predictions.csv", index=False)
# print("\nSaved: /content/lasso_only_predictions.csv")

# ============================================================
# ELASTICNET-ONLY WRITING TIME PREDICTION
# richer time features + train-only feature selection + leakage control
# ============================================================

import warnings
warnings.filterwarnings("ignore")

import numpy as np
import pandas as pd
from openpyxl import load_workbook

from sklearn.model_selection import train_test_split, KFold, GridSearchCV
from sklearn.impute import SimpleImputer
from sklearn.pipeline import Pipeline
from sklearn.preprocessing import StandardScaler
from sklearn.metrics import mean_absolute_error, mean_squared_error, r2_score
from sklearn.linear_model import ElasticNet

# ============================================================
# 1. FILE SETTINGS
# ============================================================
FILE_PATH = "<insert file path here>"
SHEET1 = "Sheet1"
SHEET2 = "Sheet2"

# ============================================================
# 2. EXCEL LAYOUT
# ============================================================
S1_START_COL = 2
S1_END_COL = 107
S1_FEATURE_ROW_START = 3
S1_FEATURE_ROW_END = 34

S2_WORD_COL = 1
S2_ORTHO_COL = 2
S2_AVGTIME_COL = 3
S2_CODES_START_COL = 4
S2_CODES_END_COL = 109
S2_ROW_START = 1
S2_ROW_END = 60

# ============================================================
# 3. LOAD WORKBOOK
# ============================================================
wb = load_workbook(FILE_PATH, data_only=True)
ws1 = wb[SHEET1]
ws2 = wb[SHEET2]

# ============================================================
# 4. READ SHEET1
# ============================================================
sheet1_rows = []
for r in range(S1_FEATURE_ROW_START, S1_FEATURE_ROW_END + 1):
    vals = []
    for c in range(S1_START_COL, S1_END_COL + 1):
        vals.append(ws1.cell(row=r, column=c).value)
    sheet1_rows.append(vals)

sheet1_rows = np.array(sheet1_rows, dtype=float)   # (32, 106)
X_sheet1 = sheet1_rows.T                           # (106, 32)

T_raw = X_sheet1[:, :30]
age = X_sheet1[:, 30]
habit = X_sheet1[:, 31]

print("Sheet1 loaded")
print("X_sheet1 shape:", X_sheet1.shape)
print("T_raw shape:", T_raw.shape)
print("age shape:", age.shape)
print("habit shape:", habit.shape)

# ============================================================
# 5. READ SHEET2
# ============================================================
words = []
ortho = []
avg_word_time = []
code_rows = []

for r in range(S2_ROW_START, S2_ROW_END + 1):
    word = ws2.cell(row=r, column=S2_WORD_COL).value
    osyll = ws2.cell(row=r, column=S2_ORTHO_COL).value
    tsec = ws2.cell(row=r, column=S2_AVGTIME_COL).value

    if word is not None and osyll is not None and tsec is not None:
        words.append(str(word))
        ortho.append(float(osyll))
        avg_word_time.append(float(tsec))

        row_codes = []
        for c in range(S2_CODES_START_COL, S2_CODES_END_COL + 1):
            row_codes.append(ws2.cell(row=r, column=c).value)
        code_rows.append(row_codes)

words = words[:30]
ortho = np.array(ortho[:30], dtype=float)
avg_word_time = np.array(avg_word_time[:30], dtype=float)
code_rows = np.array(code_rows[:30], dtype=float)

assert len(words) == 30, f"Expected 30 words, got {len(words)}"
assert code_rows.shape == (30, 106), f"Expected (30,106), got {code_rows.shape}"

codes_pw = code_rows.T  # (106, 30)

print("\nSheet2 loaded")
print("Number of words:", len(words))
print("ortho shape:", ortho.shape)
print("avg_word_time shape:", avg_word_time.shape)
print("codes_pw shape:", codes_pw.shape)

# ============================================================
# 6. HELPERS
# ============================================================
eps = 1e-8

def code_to_attempt(code):
    try:
        code = int(code)
    except:
        return np.nan

    if code in [1, 11]:
        return 1
    elif code in [2, 12]:
        return 2
    elif code in [3, 13]:
        return 3
    elif code in [4, 14]:
        return 4
    return np.nan

def safe_corr(a, b):
    a = np.asarray(a, dtype=float)
    b = np.asarray(b, dtype=float)
    if np.std(a) < 1e-12 or np.std(b) < 1e-12:
        return 0.0
    return np.corrcoef(a, b)[0, 1]

def linear_slope(x, y):
    x = np.asarray(x, dtype=float)
    y = np.asarray(y, dtype=float)
    if np.std(y) < 1e-12:
        return 0.0
    return np.polyfit(x, y, 1)[0]

def compute_train_amnesia_features(codes_pw_subset):
    n_participants, n_words = codes_pw_subset.shape

    am1_train = np.zeros(n_participants, dtype=float)
    amall_train = np.zeros(n_participants, dtype=float)
    mean_attempt_train = np.zeros(n_participants, dtype=float)
    max_attempt_train = np.zeros(n_participants, dtype=float)
    fail_rate_train = np.zeros(n_participants, dtype=float)
    delayed_rate_train = np.zeros(n_participants, dtype=float)

    for p in range(n_participants):
        row = codes_pw_subset[p]

        count_first_success = np.sum(row == 1)
        am1_train[p] = (n_words - count_first_success) / n_words

        count_total_fail = np.sum(np.isin(row, [11, 12, 13, 14]))
        amall_train[p] = count_total_fail / n_words

        attempts = np.array([code_to_attempt(x) for x in row], dtype=float)
        if np.all(np.isnan(attempts)):
            mean_attempt_train[p] = 0.0
            max_attempt_train[p] = 0.0
        else:
            mean_attempt_train[p] = np.nanmean(attempts)
            max_attempt_train[p] = np.nanmax(attempts)

        fail_rate_train[p] = count_total_fail / n_words
        delayed_rate_train[p] = np.sum(np.isin(row, [2, 3, 4])) / n_words

    return (
        am1_train,
        amall_train,
        mean_attempt_train,
        max_attempt_train,
        fail_rate_train,
        delayed_rate_train
    )

# ============================================================
# 7. WORD SPLIT
# ============================================================
rng = np.random.RandomState(42)
word_idx = np.arange(30)
rng.shuffle(word_idx)

train_word_idx = word_idx[:24]
test_word_idx = word_idx[24:30]

print("\nWord split")
print("Train words:", len(train_word_idx))
print("Test words:", len(test_word_idx))

# ============================================================
# 8. TRAIN-WORD ATTEMPT FEATURES
# ============================================================
codes_train = codes_pw[:, train_word_idx]

(
    am1_train,
    amall_train,
    mean_attempt_train,
    max_attempt_train,
    fail_rate_train,
    delayed_rate_train
) = compute_train_amnesia_features(codes_train)

# ============================================================
# 9. TARGET
# ============================================================
y_pt_proxy = np.mean(T_raw[:, test_word_idx], axis=1)

# ============================================================
# 10. PARTICIPANT SPLIT FIRST
# ============================================================
participant_ids = np.arange(1, 107)

train_idx, test_idx = train_test_split(
    np.arange(106),
    test_size=20,
    random_state=42
)

# training-only reference average to reduce leakage
avg_word_time_train_ref = np.mean(T_raw[train_idx], axis=0)

# ============================================================
# 11. FEATURE ENGINEERING
# ============================================================
def build_participant_features(
    T_raw, age, habit,
    am1_train, amall_train,
    mean_attempt_train, max_attempt_train, fail_rate_train, delayed_rate_train,
    ortho, avg_word_time_ref, train_word_idx, codes_pw
):
    Ttr = T_raw[:, train_word_idx]
    ortho_tr = ortho[train_word_idx]
    avg_time_tr = avg_word_time_ref[train_word_idx]

    n = Ttr.shape[0]
    idx = np.arange(1, Ttr.shape[1] + 1)

    attempts_tr = np.array([
        [code_to_attempt(x) for x in row]
        for row in codes_pw[:, train_word_idx]
    ], dtype=float)
    attempts_tr = np.nan_to_num(attempts_tr, nan=0.0)

    mean_t = np.mean(Ttr, axis=1)
    std_t = np.std(Ttr, axis=1)
    median_t = np.median(Ttr, axis=1)
    min_t = np.min(Ttr, axis=1)
    max_t = np.max(Ttr, axis=1)
    range_t = max_t - min_t
    cv_t = std_t / (mean_t + eps)

    slow_ratio = np.mean(Ttr > median_t[:, None], axis=1)
    extreme_ratio = np.mean(Ttr > (mean_t[:, None] + 2 * std_t[:, None]), axis=1)
    percentile_gap = np.percentile(Ttr, 90, axis=1) - np.percentile(Ttr, 10, axis=1)

    first8 = np.mean(Ttr[:, :8], axis=1)
    middle8 = np.mean(Ttr[:, 8:16], axis=1)
    last8 = np.mean(Ttr[:, 16:24], axis=1)
    late_minus_early = last8 - first8
    slope_trial = np.array([linear_slope(idx, Ttr[i]) for i in range(n)])

    norm_t = Ttr / (avg_time_tr.reshape(1, -1) + eps)
    mean_norm = np.mean(norm_t, axis=1)
    std_norm = np.std(norm_t, axis=1)
    max_norm = np.max(norm_t, axis=1)

    resid_t = Ttr - avg_time_tr.reshape(1, -1)
    mean_resid = np.mean(resid_t, axis=1)
    std_resid = np.std(resid_t, axis=1)

    weighted_ortho_time = np.sum(Ttr * ortho_tr.reshape(1, -1), axis=1) / (np.sum(ortho_tr) + eps)
    corr_time_ortho = np.array([safe_corr(Ttr[i], ortho_tr) for i in range(n)])
    corr_time_avg = np.array([safe_corr(Ttr[i], avg_time_tr) for i in range(n)])
    slope_ortho = np.array([linear_slope(ortho_tr, Ttr[i]) for i in range(n)])

    q1 = np.quantile(ortho_tr, 1/3)
    q2 = np.quantile(ortho_tr, 2/3)

    low_mask = ortho_tr <= q1
    mid_mask = (ortho_tr > q1) & (ortho_tr <= q2)
    high_mask = ortho_tr > q2

    mean_low = np.mean(Ttr[:, low_mask], axis=1)
    mean_mid = np.mean(Ttr[:, mid_mask], axis=1)
    mean_high = np.mean(Ttr[:, high_mask], axis=1)
    high_minus_low = mean_high - mean_low
    high_over_low = mean_high / (mean_low + eps)

    time_per_habit = mean_t / (habit + eps)
    age_time_interaction = age * mean_t
    habit_time_interaction = habit * mean_t

    df_feat = pd.DataFrame({
        "age": age,
        "habit": habit,

        "am1_train": am1_train,
        "amall_train": amall_train,
        "mean_attempt_train": mean_attempt_train,
        "max_attempt_train": max_attempt_train,
        "fail_rate_train": fail_rate_train,
        "delayed_rate_train": delayed_rate_train,

        "mean_t": mean_t,
        "std_t": std_t,
        "median_t": median_t,
        "min_t": min_t,
        "max_t": max_t,
        "range_t": range_t,
        "cv_t": cv_t,

        "slow_ratio": slow_ratio,
        "extreme_ratio": extreme_ratio,
        "percentile_gap": percentile_gap,

        "first8_t": first8,
        "middle8_t": middle8,
        "last8_t": last8,
        "late_minus_early": late_minus_early,
        "slope_trial": slope_trial,

        "mean_norm": mean_norm,
        "std_norm": std_norm,
        "max_norm": max_norm,
        "mean_resid": mean_resid,
        "std_resid": std_resid,

        "weighted_ortho_time": weighted_ortho_time,
        "corr_time_ortho": corr_time_ortho,
        "corr_time_avg": corr_time_avg,
        "slope_ortho": slope_ortho,

        "mean_low_ortho_time": mean_low,
        "mean_mid_ortho_time": mean_mid,
        "mean_high_ortho_time": mean_high,
        "high_minus_low": high_minus_low,
        "high_over_low": high_over_low,

        "time_per_habit": time_per_habit,
        "age_time_interaction": age_time_interaction,
        "habit_time_interaction": habit_time_interaction
    })

    for j in range(Ttr.shape[1]):
        df_feat[f"raw_time_{j+1:02d}"] = Ttr[:, j]

    for j in range(norm_t.shape[1]):
        df_feat[f"norm_time_{j+1:02d}"] = norm_t[:, j]

    for j in range(attempts_tr.shape[1]):
        df_feat[f"attempt_{j+1:02d}"] = attempts_tr[:, j]

    return df_feat.fillna(0)

X_feat = build_participant_features(
    T_raw, age, habit,
    am1_train, amall_train,
    mean_attempt_train, max_attempt_train, fail_rate_train, delayed_rate_train,
    ortho, avg_word_time_train_ref, train_word_idx, codes_pw
)

# ============================================================
# 12. TRAIN / TEST MATRICES
# ============================================================
X_train_full = X_feat.iloc[train_idx].reset_index(drop=True)
X_test_full = X_feat.iloc[test_idx].reset_index(drop=True)

y_train = y_pt_proxy[train_idx]
y_test = y_pt_proxy[test_idx]

id_test = participant_ids[test_idx]

# ============================================================
# 13. TRAIN-ONLY FEATURE SELECTION
# ============================================================
def select_top_features(X_df, y, top_k=30):
    corrs = []
    for col in X_df.columns:
        c = safe_corr(X_df[col].values, y)
        corrs.append(abs(c))
    corrs = np.array(corrs)
    idx = np.argsort(corrs)[-top_k:]
    cols = list(X_df.columns[idx])
    return cols, pd.Series(corrs, index=X_df.columns).sort_values(ascending=False)

TOP_K = 30
selected_cols, corr_series = select_top_features(X_train_full, y_train, top_k=TOP_K)

X_train = X_train_full[selected_cols].copy()
X_test = X_test_full[selected_cols].copy()

print("\nSelected feature matrix shape (train):", X_train.shape)
print("Selected feature matrix shape (test):", X_test.shape)
print("\nTop correlated training features:")
print(corr_series.head(15))

# ============================================================
# 14. XGBOOST ONLY
# ============================================================
enet_pipe = Pipeline([
    ("imputer", SimpleImputer(strategy="median")),
    ("scaler", StandardScaler()),
    ("model", ElasticNet())
])

enet_param_grid = {}

cv = KFold(n_splits=5, shuffle=True, random_state=42)

search = GridSearchCV(
    estimator=enet_pipe,
    param_grid=enet_param_grid,
    scoring="r2",
    cv=cv,
    n_jobs=-1,
    refit=True,
    error_score="raise"
)

print("\n==================== ELASTICNET ONLY ====================")
search.fit(X_train, y_train)

y_pred = search.best_estimator_.predict(X_test)

mae = mean_absolute_error(y_test, y_pred)
mse = mean_squared_error(y_test, y_pred)
rmse = np.sqrt(mse)
r2 = r2_score(y_test, y_pred)

print("\nBest Params:")
print(search.best_params_)

print("\nMetrics:")
print(f"BestCV_R2 = {search.best_score_:.4f}")
print(f"Test_MAE  = {mae:.4f}")
print(f"Test_MSE  = {mse:.4f}")
print(f"Test_RMSE = {rmse:.4f}")
print(f"Test_R2   = {r2:.4f}")

pred_df = pd.DataFrame({
    "participant_id": id_test,
    "true_avg_time_test6": y_test,
    "pred_avg_time_test6": y_pred
})

print("\nPredictions preview:")
print(pred_df.head())

# Optional save
# pred_df.to_csv("/content/elasticnet_only_predictions.csv", index=False)
# print("\nSaved: /content/elasticnet_only_predictions.csv")

# ============================================================
# SVR-ONLY WRITING TIME PREDICTION
# ============================================================

import warnings
warnings.filterwarnings("ignore")

import numpy as np
import pandas as pd
from openpyxl import load_workbook

from sklearn.model_selection import train_test_split, KFold, GridSearchCV
from sklearn.impute import SimpleImputer
from sklearn.pipeline import Pipeline
from sklearn.preprocessing import StandardScaler
from sklearn.metrics import mean_absolute_error, mean_squared_error, r2_score
from sklearn.svm import SVR

# ============================================================
# 1. FILE SETTINGS
# ============================================================
FILE_PATH = "<insert file path here>"
SHEET1 = "Sheet1"
SHEET2 = "Sheet2"

# ============================================================
# 2. EXCEL LAYOUT
# ============================================================
S1_START_COL = 2
S1_END_COL = 107
S1_FEATURE_ROW_START = 3
S1_FEATURE_ROW_END = 34

S2_WORD_COL = 1
S2_ORTHO_COL = 2
S2_AVGTIME_COL = 3
S2_CODES_START_COL = 4
S2_CODES_END_COL = 109
S2_ROW_START = 1
S2_ROW_END = 60

# ============================================================
# 3. LOAD WORKBOOK
# ============================================================
wb = load_workbook(FILE_PATH, data_only=True)
ws1 = wb[SHEET1]
ws2 = wb[SHEET2]

# ============================================================
# 4. READ SHEET1
# ============================================================
sheet1_rows = []
for r in range(S1_FEATURE_ROW_START, S1_FEATURE_ROW_END + 1):
    vals = []
    for c in range(S1_START_COL, S1_END_COL + 1):
        vals.append(ws1.cell(row=r, column=c).value)
    sheet1_rows.append(vals)

sheet1_rows = np.array(sheet1_rows, dtype=float)   # (32, 106)
X_sheet1 = sheet1_rows.T                           # (106, 32)

T_raw = X_sheet1[:, :30]
age = X_sheet1[:, 30]
habit = X_sheet1[:, 31]

print("Sheet1 loaded")
print("X_sheet1 shape:", X_sheet1.shape)
print("T_raw shape:", T_raw.shape)
print("age shape:", age.shape)
print("habit shape:", habit.shape)

# ============================================================
# 5. READ SHEET2
# ============================================================
words = []
ortho = []
avg_word_time = []
code_rows = []

for r in range(S2_ROW_START, S2_ROW_END + 1):
    word = ws2.cell(row=r, column=S2_WORD_COL).value
    osyll = ws2.cell(row=r, column=S2_ORTHO_COL).value
    tsec = ws2.cell(row=r, column=S2_AVGTIME_COL).value

    if word is not None and osyll is not None and tsec is not None:
        words.append(str(word))
        ortho.append(float(osyll))
        avg_word_time.append(float(tsec))

        row_codes = []
        for c in range(S2_CODES_START_COL, S2_CODES_END_COL + 1):
            row_codes.append(ws2.cell(row=r, column=c).value)
        code_rows.append(row_codes)

words = words[:30]
ortho = np.array(ortho[:30], dtype=float)
avg_word_time = np.array(avg_word_time[:30], dtype=float)
code_rows = np.array(code_rows[:30], dtype=float)

assert len(words) == 30, f"Expected 30 words, got {len(words)}"
assert code_rows.shape == (30, 106), f"Expected (30,106), got {code_rows.shape}"

codes_pw = code_rows.T  # (106, 30)

print("\nSheet2 loaded")
print("Number of words:", len(words))
print("ortho shape:", ortho.shape)
print("avg_word_time shape:", avg_word_time.shape)
print("codes_pw shape:", codes_pw.shape)

# ============================================================
# 6. HELPERS
# ============================================================
eps = 1e-8

def code_to_attempt(code):
    try:
        code = int(code)
    except:
        return np.nan

    if code in [1, 11]:
        return 1
    elif code in [2, 12]:
        return 2
    elif code in [3, 13]:
        return 3
    elif code in [4, 14]:
        return 4
    return np.nan

def safe_corr(a, b):
    a = np.asarray(a, dtype=float)
    b = np.asarray(b, dtype=float)
    if np.std(a) < 1e-12 or np.std(b) < 1e-12:
        return 0.0
    return np.corrcoef(a, b)[0, 1]

def linear_slope(x, y):
    x = np.asarray(x, dtype=float)
    y = np.asarray(y, dtype=float)
    if np.std(y) < 1e-12:
        return 0.0
    return np.polyfit(x, y, 1)[0]

def compute_train_amnesia_features(codes_pw_subset):
    n_participants, n_words = codes_pw_subset.shape

    am1_train = np.zeros(n_participants, dtype=float)
    amall_train = np.zeros(n_participants, dtype=float)
    mean_attempt_train = np.zeros(n_participants, dtype=float)
    max_attempt_train = np.zeros(n_participants, dtype=float)
    fail_rate_train = np.zeros(n_participants, dtype=float)
    delayed_rate_train = np.zeros(n_participants, dtype=float)

    for p in range(n_participants):
        row = codes_pw_subset[p]

        count_first_success = np.sum(row == 1)
        am1_train[p] = (n_words - count_first_success) / n_words

        count_total_fail = np.sum(np.isin(row, [11, 12, 13, 14]))
        amall_train[p] = count_total_fail / n_words

        attempts = np.array([code_to_attempt(x) for x in row], dtype=float)
        if np.all(np.isnan(attempts)):
            mean_attempt_train[p] = 0.0
            max_attempt_train[p] = 0.0
        else:
            mean_attempt_train[p] = np.nanmean(attempts)
            max_attempt_train[p] = np.nanmax(attempts)

        fail_rate_train[p] = count_total_fail / n_words
        delayed_rate_train[p] = np.sum(np.isin(row, [2, 3, 4])) / n_words

    return (
        am1_train,
        amall_train,
        mean_attempt_train,
        max_attempt_train,
        fail_rate_train,
        delayed_rate_train
    )

# ============================================================
# 7. WORD SPLIT
# ============================================================
rng = np.random.RandomState(42)
word_idx = np.arange(30)
rng.shuffle(word_idx)

train_word_idx = word_idx[:24]
test_word_idx = word_idx[24:30]

print("\nWord split")
print("Train words:", len(train_word_idx))
print("Test words:", len(test_word_idx))

# ============================================================
# 8. TRAIN-WORD ATTEMPT FEATURES
# ============================================================
codes_train = codes_pw[:, train_word_idx]

(
    am1_train,
    amall_train,
    mean_attempt_train,
    max_attempt_train,
    fail_rate_train,
    delayed_rate_train
) = compute_train_amnesia_features(codes_train)

# ============================================================
# 9. TARGET
# ============================================================
y_pt_proxy = np.mean(T_raw[:, test_word_idx], axis=1)

# ============================================================
# 10. PARTICIPANT SPLIT FIRST
# ============================================================
participant_ids = np.arange(1, 107)

train_idx, test_idx = train_test_split(
    np.arange(106),
    test_size=20,
    random_state=42
)

# training-only reference average to reduce leakage
avg_word_time_train_ref = np.mean(T_raw[train_idx], axis=0)

# ============================================================
# 11. FEATURE ENGINEERING
# ============================================================
def build_participant_features(
    T_raw, age, habit,
    am1_train, amall_train,
    mean_attempt_train, max_attempt_train, fail_rate_train, delayed_rate_train,
    ortho, avg_word_time_ref, train_word_idx, codes_pw
):
    Ttr = T_raw[:, train_word_idx]
    ortho_tr = ortho[train_word_idx]
    avg_time_tr = avg_word_time_ref[train_word_idx]

    n = Ttr.shape[0]
    idx = np.arange(1, Ttr.shape[1] + 1)

    attempts_tr = np.array([
        [code_to_attempt(x) for x in row]
        for row in codes_pw[:, train_word_idx]
    ], dtype=float)
    attempts_tr = np.nan_to_num(attempts_tr, nan=0.0)

    mean_t = np.mean(Ttr, axis=1)
    std_t = np.std(Ttr, axis=1)
    median_t = np.median(Ttr, axis=1)
    min_t = np.min(Ttr, axis=1)
    max_t = np.max(Ttr, axis=1)
    range_t = max_t - min_t
    cv_t = std_t / (mean_t + eps)

    slow_ratio = np.mean(Ttr > median_t[:, None], axis=1)
    extreme_ratio = np.mean(Ttr > (mean_t[:, None] + 2 * std_t[:, None]), axis=1)
    percentile_gap = np.percentile(Ttr, 90, axis=1) - np.percentile(Ttr, 10, axis=1)

    first8 = np.mean(Ttr[:, :8], axis=1)
    middle8 = np.mean(Ttr[:, 8:16], axis=1)
    last8 = np.mean(Ttr[:, 16:24], axis=1)
    late_minus_early = last8 - first8
    slope_trial = np.array([linear_slope(idx, Ttr[i]) for i in range(n)])

    norm_t = Ttr / (avg_time_tr.reshape(1, -1) + eps)
    mean_norm = np.mean(norm_t, axis=1)
    std_norm = np.std(norm_t, axis=1)
    max_norm = np.max(norm_t, axis=1)

    resid_t = Ttr - avg_time_tr.reshape(1, -1)
    mean_resid = np.mean(resid_t, axis=1)
    std_resid = np.std(resid_t, axis=1)

    weighted_ortho_time = np.sum(Ttr * ortho_tr.reshape(1, -1), axis=1) / (np.sum(ortho_tr) + eps)
    corr_time_ortho = np.array([safe_corr(Ttr[i], ortho_tr) for i in range(n)])
    corr_time_avg = np.array([safe_corr(Ttr[i], avg_time_tr) for i in range(n)])
    slope_ortho = np.array([linear_slope(ortho_tr, Ttr[i]) for i in range(n)])

    q1 = np.quantile(ortho_tr, 1/3)
    q2 = np.quantile(ortho_tr, 2/3)

    low_mask = ortho_tr <= q1
    mid_mask = (ortho_tr > q1) & (ortho_tr <= q2)
    high_mask = ortho_tr > q2

    mean_low = np.mean(Ttr[:, low_mask], axis=1)
    mean_mid = np.mean(Ttr[:, mid_mask], axis=1)
    mean_high = np.mean(Ttr[:, high_mask], axis=1)
    high_minus_low = mean_high - mean_low
    high_over_low = mean_high / (mean_low + eps)

    time_per_habit = mean_t / (habit + eps)
    age_time_interaction = age * mean_t
    habit_time_interaction = habit * mean_t

    df_feat = pd.DataFrame({
        "age": age,
        "habit": habit,

        "am1_train": am1_train,
        "amall_train": amall_train,
        "mean_attempt_train": mean_attempt_train,
        "max_attempt_train": max_attempt_train,
        "fail_rate_train": fail_rate_train,
        "delayed_rate_train": delayed_rate_train,

        "mean_t": mean_t,
        "std_t": std_t,
        "median_t": median_t,
        "min_t": min_t,
        "max_t": max_t,
        "range_t": range_t,
        "cv_t": cv_t,

        "slow_ratio": slow_ratio,
        "extreme_ratio": extreme_ratio,
        "percentile_gap": percentile_gap,

        "first8_t": first8,
        "middle8_t": middle8,
        "last8_t": last8,
        "late_minus_early": late_minus_early,
        "slope_trial": slope_trial,

        "mean_norm": mean_norm,
        "std_norm": std_norm,
        "max_norm": max_norm,
        "mean_resid": mean_resid,
        "std_resid": std_resid,

        "weighted_ortho_time": weighted_ortho_time,
        "corr_time_ortho": corr_time_ortho,
        "corr_time_avg": corr_time_avg,
        "slope_ortho": slope_ortho,

        "mean_low_ortho_time": mean_low,
        "mean_mid_ortho_time": mean_mid,
        "mean_high_ortho_time": mean_high,
        "high_minus_low": high_minus_low,
        "high_over_low": high_over_low,

        "time_per_habit": time_per_habit,
        "age_time_interaction": age_time_interaction,
        "habit_time_interaction": habit_time_interaction
    })

    for j in range(Ttr.shape[1]):
        df_feat[f"raw_time_{j+1:02d}"] = Ttr[:, j]

    for j in range(norm_t.shape[1]):
        df_feat[f"norm_time_{j+1:02d}"] = norm_t[:, j]

    for j in range(attempts_tr.shape[1]):
        df_feat[f"attempt_{j+1:02d}"] = attempts_tr[:, j]

    return df_feat.fillna(0)

X_feat = build_participant_features(
    T_raw, age, habit,
    am1_train, amall_train,
    mean_attempt_train, max_attempt_train, fail_rate_train, delayed_rate_train,
    ortho, avg_word_time_train_ref, train_word_idx, codes_pw
)

# ============================================================
# 12. TRAIN / TEST MATRICES
# ============================================================
X_train_full = X_feat.iloc[train_idx].reset_index(drop=True)
X_test_full = X_feat.iloc[test_idx].reset_index(drop=True)

y_train = y_pt_proxy[train_idx]
y_test = y_pt_proxy[test_idx]

id_test = participant_ids[test_idx]

# ============================================================
# 13. TRAIN-ONLY FEATURE SELECTION
# ============================================================
def select_top_features(X_df, y, top_k=40):
    corrs = []
    for col in X_df.columns:
        c = safe_corr(X_df[col].values, y)
        corrs.append(abs(c))
    corrs = np.array(corrs)
    idx = np.argsort(corrs)[-top_k:]
    cols = list(X_df.columns[idx])
    return cols, pd.Series(corrs, index=X_df.columns).sort_values(ascending=False)

TOP_K = 40
selected_cols, corr_series = select_top_features(X_train_full, y_train, top_k=TOP_K)

X_train = X_train_full[selected_cols].copy()
X_test = X_test_full[selected_cols].copy()

print("\nSelected feature matrix shape (train):", X_train.shape)
print("Selected feature matrix shape (test):", X_test.shape)
print("\nTop correlated training features:")
print(corr_series.head(15))

# ============================================================
# 14. SVR ONLY
# ============================================================
svr_pipe = Pipeline([
    ("imputer", SimpleImputer(strategy="median")),
    ("scaler", StandardScaler()),
    ("model", SVR())
])

svr_param_grid = {}

cv = KFold(n_splits=5, shuffle=True, random_state=42)

search = GridSearchCV(
    estimator=svr_pipe,
    param_grid=svr_param_grid,
    scoring="r2",
    cv=cv,
    n_jobs=-1,
    refit=True,
    error_score="raise"
)

print("\n==================== SVR ONLY ====================")
search.fit(X_train, y_train)

y_pred = search.best_estimator_.predict(X_test)

mae = mean_absolute_error(y_test, y_pred)
mse = mean_squared_error(y_test, y_pred)
rmse = np.sqrt(mse)
r2 = r2_score(y_test, y_pred)

print("\nBest Params:")
print(search.best_params_)

print("\nMetrics:")
print(f"BestCV_R2 = {search.best_score_:.4f}")
print(f"Test_MAE  = {mae:.4f}")
print(f"Test_MSE  = {mse:.4f}")
print(f"Test_RMSE = {rmse:.4f}")
print(f"Test_R2   = {r2:.4f}")

pred_df = pd.DataFrame({
    "participant_id": id_test,
    "true_avg_time_test6": y_test,
    "pred_avg_time_test6": y_pred
})

print("\nPredictions preview:")
print(pred_df.head())

# Optional save
# pred_df.to_csv("/content/svr_only_predictions.csv", index=False)
# print("\nSaved: /content/svr_only_predictions.csv")

# ============================================================
# KNNR-ONLY WRITING TIME PREDICTION
# ============================================================

import warnings
warnings.filterwarnings("ignore")

import numpy as np
import pandas as pd
from openpyxl import load_workbook

from sklearn.model_selection import train_test_split, KFold, GridSearchCV
from sklearn.impute import SimpleImputer
from sklearn.pipeline import Pipeline
from sklearn.preprocessing import StandardScaler
from sklearn.metrics import mean_absolute_error, mean_squared_error, r2_score
from sklearn.neighbors import KNeighborsRegressor

# ============================================================
# 1. FILE SETTINGS
# ============================================================
FILE_PATH ="<insert file path here>"
SHEET1 = "Sheet1"
SHEET2 = "Sheet2"

# ============================================================
# 2. EXCEL LAYOUT
# ============================================================
S1_START_COL = 2
S1_END_COL = 107
S1_FEATURE_ROW_START = 3
S1_FEATURE_ROW_END = 34

S2_WORD_COL = 1
S2_ORTHO_COL = 2
S2_AVGTIME_COL = 3
S2_CODES_START_COL = 4
S2_CODES_END_COL = 109
S2_ROW_START = 1
S2_ROW_END = 60

# ============================================================
# 3. LOAD WORKBOOK
# ============================================================
wb = load_workbook(FILE_PATH, data_only=True)
ws1 = wb[SHEET1]
ws2 = wb[SHEET2]

# ============================================================
# 4. READ SHEET1
# ============================================================
sheet1_rows = []
for r in range(S1_FEATURE_ROW_START, S1_FEATURE_ROW_END + 1):
    vals = []
    for c in range(S1_START_COL, S1_END_COL + 1):
        vals.append(ws1.cell(row=r, column=c).value)
    sheet1_rows.append(vals)

sheet1_rows = np.array(sheet1_rows, dtype=float)   # (32, 106)
X_sheet1 = sheet1_rows.T                           # (106, 32)

T_raw = X_sheet1[:, :30]
age = X_sheet1[:, 30]
habit = X_sheet1[:, 31]

print("Sheet1 loaded")
print("X_sheet1 shape:", X_sheet1.shape)
print("T_raw shape:", T_raw.shape)
print("age shape:", age.shape)
print("habit shape:", habit.shape)

# ============================================================
# 5. READ SHEET2
# ============================================================
words = []
ortho = []
avg_word_time = []
code_rows = []

for r in range(S2_ROW_START, S2_ROW_END + 1):
    word = ws2.cell(row=r, column=S2_WORD_COL).value
    osyll = ws2.cell(row=r, column=S2_ORTHO_COL).value
    tsec = ws2.cell(row=r, column=S2_AVGTIME_COL).value

    if word is not None and osyll is not None and tsec is not None:
        words.append(str(word))
        ortho.append(float(osyll))
        avg_word_time.append(float(tsec))

        row_codes = []
        for c in range(S2_CODES_START_COL, S2_CODES_END_COL + 1):
            row_codes.append(ws2.cell(row=r, column=c).value)
        code_rows.append(row_codes)

words = words[:30]
ortho = np.array(ortho[:30], dtype=float)
avg_word_time = np.array(avg_word_time[:30], dtype=float)
code_rows = np.array(code_rows[:30], dtype=float)

assert len(words) == 30, f"Expected 30 words, got {len(words)}"
assert code_rows.shape == (30, 106), f"Expected (30,106), got {code_rows.shape}"

codes_pw = code_rows.T  # (106, 30)

print("\nSheet2 loaded")
print("Number of words:", len(words))
print("ortho shape:", ortho.shape)
print("avg_word_time shape:", avg_word_time.shape)
print("codes_pw shape:", codes_pw.shape)

# ============================================================
# 6. HELPERS
# ============================================================
eps = 1e-8

def code_to_attempt(code):
    try:
        code = int(code)
    except:
        return np.nan

    if code in [1, 11]:
        return 1
    elif code in [2, 12]:
        return 2
    elif code in [3, 13]:
        return 3
    elif code in [4, 14]:
        return 4
    return np.nan

def safe_corr(a, b):
    a = np.asarray(a, dtype=float)
    b = np.asarray(b, dtype=float)
    if np.std(a) < 1e-12 or np.std(b) < 1e-12:
        return 0.0
    return np.corrcoef(a, b)[0, 1]

def linear_slope(x, y):
    x = np.asarray(x, dtype=float)
    y = np.asarray(y, dtype=float)
    if np.std(y) < 1e-12:
        return 0.0
    return np.polyfit(x, y, 1)[0]

def compute_train_amnesia_features(codes_pw_subset):
    n_participants, n_words = codes_pw_subset.shape

    am1_train = np.zeros(n_participants, dtype=float)
    amall_train = np.zeros(n_participants, dtype=float)
    mean_attempt_train = np.zeros(n_participants, dtype=float)
    max_attempt_train = np.zeros(n_participants, dtype=float)
    fail_rate_train = np.zeros(n_participants, dtype=float)
    delayed_rate_train = np.zeros(n_participants, dtype=float)

    for p in range(n_participants):
        row = codes_pw_subset[p]

        count_first_success = np.sum(row == 1)
        am1_train[p] = (n_words - count_first_success) / n_words

        count_total_fail = np.sum(np.isin(row, [11, 12, 13, 14]))
        amall_train[p] = count_total_fail / n_words

        attempts = np.array([code_to_attempt(x) for x in row], dtype=float)
        if np.all(np.isnan(attempts)):
            mean_attempt_train[p] = 0.0
            max_attempt_train[p] = 0.0
        else:
            mean_attempt_train[p] = np.nanmean(attempts)
            max_attempt_train[p] = np.nanmax(attempts)

        fail_rate_train[p] = count_total_fail / n_words
        delayed_rate_train[p] = np.sum(np.isin(row, [2, 3, 4])) / n_words

    return (
        am1_train,
        amall_train,
        mean_attempt_train,
        max_attempt_train,
        fail_rate_train,
        delayed_rate_train
    )

# ============================================================
# 7. WORD SPLIT
# ============================================================
rng = np.random.RandomState(42)
word_idx = np.arange(30)
rng.shuffle(word_idx)

train_word_idx = word_idx[:24]
test_word_idx = word_idx[24:30]

print("\nWord split")
print("Train words:", len(train_word_idx))
print("Test words:", len(test_word_idx))

# ============================================================
# 8. TRAIN-WORD ATTEMPT FEATURES
# ============================================================
codes_train = codes_pw[:, train_word_idx]

(
    am1_train,
    amall_train,
    mean_attempt_train,
    max_attempt_train,
    fail_rate_train,
    delayed_rate_train
) = compute_train_amnesia_features(codes_train)

# ============================================================
# 9. TARGET
# ============================================================
y_pt_proxy = np.mean(T_raw[:, test_word_idx], axis=1)

# ============================================================
# 10. PARTICIPANT SPLIT FIRST
# ============================================================
participant_ids = np.arange(1, 107)

train_idx, test_idx = train_test_split(
    np.arange(106),
    test_size=20,
    random_state=42
)

# training-only reference average to reduce leakage
avg_word_time_train_ref = np.mean(T_raw[train_idx], axis=0)

# ============================================================
# 11. FEATURE ENGINEERING
# ============================================================
def build_participant_features(
    T_raw, age, habit,
    am1_train, amall_train,
    mean_attempt_train, max_attempt_train, fail_rate_train, delayed_rate_train,
    ortho, avg_word_time_ref, train_word_idx, codes_pw
):
    Ttr = T_raw[:, train_word_idx]
    ortho_tr = ortho[train_word_idx]
    avg_time_tr = avg_word_time_ref[train_word_idx]

    n = Ttr.shape[0]
    idx = np.arange(1, Ttr.shape[1] + 1)

    attempts_tr = np.array([
        [code_to_attempt(x) for x in row]
        for row in codes_pw[:, train_word_idx]
    ], dtype=float)
    attempts_tr = np.nan_to_num(attempts_tr, nan=0.0)

    mean_t = np.mean(Ttr, axis=1)
    std_t = np.std(Ttr, axis=1)
    median_t = np.median(Ttr, axis=1)
    min_t = np.min(Ttr, axis=1)
    max_t = np.max(Ttr, axis=1)
    range_t = max_t - min_t
    cv_t = std_t / (mean_t + eps)

    slow_ratio = np.mean(Ttr > median_t[:, None], axis=1)
    extreme_ratio = np.mean(Ttr > (mean_t[:, None] + 2 * std_t[:, None]), axis=1)
    percentile_gap = np.percentile(Ttr, 90, axis=1) - np.percentile(Ttr, 10, axis=1)

    first8 = np.mean(Ttr[:, :8], axis=1)
    middle8 = np.mean(Ttr[:, 8:16], axis=1)
    last8 = np.mean(Ttr[:, 16:24], axis=1)
    late_minus_early = last8 - first8
    slope_trial = np.array([linear_slope(idx, Ttr[i]) for i in range(n)])

    norm_t = Ttr / (avg_time_tr.reshape(1, -1) + eps)
    mean_norm = np.mean(norm_t, axis=1)
    std_norm = np.std(norm_t, axis=1)
    max_norm = np.max(norm_t, axis=1)

    resid_t = Ttr - avg_time_tr.reshape(1, -1)
    mean_resid = np.mean(resid_t, axis=1)
    std_resid = np.std(resid_t, axis=1)

    weighted_ortho_time = np.sum(Ttr * ortho_tr.reshape(1, -1), axis=1) / (np.sum(ortho_tr) + eps)
    corr_time_ortho = np.array([safe_corr(Ttr[i], ortho_tr) for i in range(n)])
    corr_time_avg = np.array([safe_corr(Ttr[i], avg_time_tr) for i in range(n)])
    slope_ortho = np.array([linear_slope(ortho_tr, Ttr[i]) for i in range(n)])

    q1 = np.quantile(ortho_tr, 1/3)
    q2 = np.quantile(ortho_tr, 2/3)

    low_mask = ortho_tr <= q1
    mid_mask = (ortho_tr > q1) & (ortho_tr <= q2)
    high_mask = ortho_tr > q2

    mean_low = np.mean(Ttr[:, low_mask], axis=1)
    mean_mid = np.mean(Ttr[:, mid_mask], axis=1)
    mean_high = np.mean(Ttr[:, high_mask], axis=1)
    high_minus_low = mean_high - mean_low
    high_over_low = mean_high / (mean_low + eps)

    time_per_habit = mean_t / (habit + eps)
    age_time_interaction = age * mean_t
    habit_time_interaction = habit * mean_t

    df_feat = pd.DataFrame({
        "age": age,
        "habit": habit,

        "am1_train": am1_train,
        "amall_train": amall_train,
        "mean_attempt_train": mean_attempt_train,
        "max_attempt_train": max_attempt_train,
        "fail_rate_train": fail_rate_train,
        "delayed_rate_train": delayed_rate_train,

        "mean_t": mean_t,
        "std_t": std_t,
        "median_t": median_t,
        "min_t": min_t,
        "max_t": max_t,
        "range_t": range_t,
        "cv_t": cv_t,

        "slow_ratio": slow_ratio,
        "extreme_ratio": extreme_ratio,
        "percentile_gap": percentile_gap,

        "first8_t": first8,
        "middle8_t": middle8,
        "last8_t": last8,
        "late_minus_early": late_minus_early,
        "slope_trial": slope_trial,

        "mean_norm": mean_norm,
        "std_norm": std_norm,
        "max_norm": max_norm,
        "mean_resid": mean_resid,
        "std_resid": std_resid,

        "weighted_ortho_time": weighted_ortho_time,
        "corr_time_ortho": corr_time_ortho,
        "corr_time_avg": corr_time_avg,
        "slope_ortho": slope_ortho,

        "mean_low_ortho_time": mean_low,
        "mean_mid_ortho_time": mean_mid,
        "mean_high_ortho_time": mean_high,
        "high_minus_low": high_minus_low,
        "high_over_low": high_over_low,

        "time_per_habit": time_per_habit,
        "age_time_interaction": age_time_interaction,
        "habit_time_interaction": habit_time_interaction
    })

    for j in range(Ttr.shape[1]):
        df_feat[f"raw_time_{j+1:02d}"] = Ttr[:, j]

    for j in range(norm_t.shape[1]):
        df_feat[f"norm_time_{j+1:02d}"] = norm_t[:, j]

    for j in range(attempts_tr.shape[1]):
        df_feat[f"attempt_{j+1:02d}"] = attempts_tr[:, j]

    return df_feat.fillna(0)

X_feat = build_participant_features(
    T_raw, age, habit,
    am1_train, amall_train,
    mean_attempt_train, max_attempt_train, fail_rate_train, delayed_rate_train,
    ortho, avg_word_time_train_ref, train_word_idx, codes_pw
)

# ============================================================
# 12. TRAIN / TEST MATRICES
# ============================================================
X_train_full = X_feat.iloc[train_idx].reset_index(drop=True)
X_test_full = X_feat.iloc[test_idx].reset_index(drop=True)

y_train = y_pt_proxy[train_idx]
y_test = y_pt_proxy[test_idx]

id_test = participant_ids[test_idx]

# ============================================================
# 13. TRAIN-ONLY FEATURE SELECTION
# ============================================================
def select_top_features(X_df, y, top_k=30):
    corrs = []
    for col in X_df.columns:
        c = safe_corr(X_df[col].values, y)
        corrs.append(abs(c))
    corrs = np.array(corrs)
    idx = np.argsort(corrs)[-top_k:]
    cols = list(X_df.columns[idx])
    return cols, pd.Series(corrs, index=X_df.columns).sort_values(ascending=False)

TOP_K = 30
selected_cols, corr_series = select_top_features(X_train_full, y_train, top_k=TOP_K)

X_train = X_train_full[selected_cols].copy()
X_test = X_test_full[selected_cols].copy()

print("\nSelected feature matrix shape (train):", X_train.shape)
print("Selected feature matrix shape (test):", X_test.shape)
print("\nTop correlated training features:")
print(corr_series.head(15))

# ============================================================
# 14. KNNR ONLY
# ============================================================
knnr_pipe = Pipeline([
    ("imputer", SimpleImputer(strategy="median")),
    ("scaler", StandardScaler()),
    ("model", KNeighborsRegressor())
])

knnr_param_grid = {}

cv = KFold(n_splits=5, shuffle=True, random_state=42)

search = GridSearchCV(
    estimator=knnr_pipe,
    param_grid=knnr_param_grid,
    scoring="r2",
    cv=cv,
    n_jobs=-1,
    refit=True,
    error_score="raise"
)

print("\n==================== KNNR ONLY ====================")
search.fit(X_train, y_train)

y_pred = search.best_estimator_.predict(X_test)

mae = mean_absolute_error(y_test, y_pred)
mse = mean_squared_error(y_test, y_pred)
rmse = np.sqrt(mse)
r2 = r2_score(y_test, y_pred)

print("\nBest Params:")
print(search.best_params_)

print("\nMetrics:")
print(f"BestCV_R2 = {search.best_score_:.4f}")
print(f"Test_MAE  = {mae:.4f}")
print(f"Test_MSE  = {mse:.4f}")
print(f"Test_RMSE = {rmse:.4f}")
print(f"Test_R2   = {r2:.4f}")

pred_df = pd.DataFrame({
    "participant_id": id_test,
    "true_avg_time_test6": y_test,
    "pred_avg_time_test6": y_pred
})

print("\nPredictions preview:")
print(pred_df.head())

# Optional save
# pred_df.to_csv("/content/knnr_only_predictions.csv", index=False)
# print("\nSaved: /content/knnr_only_predictions.csv")

# ============================================================
# RFR-ONLY WRITING TIME PREDICTION
# ============================================================

import warnings
warnings.filterwarnings("ignore")

import numpy as np
import pandas as pd
from openpyxl import load_workbook

from sklearn.model_selection import train_test_split, KFold, GridSearchCV
from sklearn.impute import SimpleImputer
from sklearn.pipeline import Pipeline
from sklearn.metrics import mean_absolute_error, mean_squared_error, r2_score
from sklearn.ensemble import RandomForestRegressor

# ============================================================
# 1. FILE SETTINGS
# ============================================================
FILE_PATH = "<insert file path here>"
SHEET1 = "Sheet1"
SHEET2 = "Sheet2"

# ============================================================
# 2. EXCEL LAYOUT
# ============================================================
S1_START_COL = 2
S1_END_COL = 107
S1_FEATURE_ROW_START = 3
S1_FEATURE_ROW_END = 34

S2_WORD_COL = 1
S2_ORTHO_COL = 2
S2_AVGTIME_COL = 3
S2_CODES_START_COL = 4
S2_CODES_END_COL = 109
S2_ROW_START = 1
S2_ROW_END = 60

# ============================================================
# 3. LOAD WORKBOOK
# ============================================================
wb = load_workbook(FILE_PATH, data_only=True)
ws1 = wb[SHEET1]
ws2 = wb[SHEET2]

# ============================================================
# 4. READ SHEET1
# ============================================================
sheet1_rows = []
for r in range(S1_FEATURE_ROW_START, S1_FEATURE_ROW_END + 1):
    vals = []
    for c in range(S1_START_COL, S1_END_COL + 1):
        vals.append(ws1.cell(row=r, column=c).value)
    sheet1_rows.append(vals)

sheet1_rows = np.array(sheet1_rows, dtype=float)   # (32, 106)
X_sheet1 = sheet1_rows.T                           # (106, 32)

T_raw = X_sheet1[:, :30]
age = X_sheet1[:, 30]
habit = X_sheet1[:, 31]

print("Sheet1 loaded")
print("X_sheet1 shape:", X_sheet1.shape)
print("T_raw shape:", T_raw.shape)
print("age shape:", age.shape)
print("habit shape:", habit.shape)

# ============================================================
# 5. READ SHEET2
# ============================================================
words = []
ortho = []
avg_word_time = []
code_rows = []

for r in range(S2_ROW_START, S2_ROW_END + 1):
    word = ws2.cell(row=r, column=S2_WORD_COL).value
    osyll = ws2.cell(row=r, column=S2_ORTHO_COL).value
    tsec = ws2.cell(row=r, column=S2_AVGTIME_COL).value

    if word is not None and osyll is not None and tsec is not None:
        words.append(str(word))
        ortho.append(float(osyll))
        avg_word_time.append(float(tsec))

        row_codes = []
        for c in range(S2_CODES_START_COL, S2_CODES_END_COL + 1):
            row_codes.append(ws2.cell(row=r, column=c).value)
        code_rows.append(row_codes)

words = words[:30]
ortho = np.array(ortho[:30], dtype=float)
avg_word_time = np.array(avg_word_time[:30], dtype=float)
code_rows = np.array(code_rows[:30], dtype=float)

assert len(words) == 30, f"Expected 30 words, got {len(words)}"
assert code_rows.shape == (30, 106), f"Expected (30,106), got {code_rows.shape}"

codes_pw = code_rows.T  # (106, 30)

print("\nSheet2 loaded")
print("Number of words:", len(words))
print("ortho shape:", ortho.shape)
print("avg_word_time shape:", avg_word_time.shape)
print("codes_pw shape:", codes_pw.shape)

# ============================================================
# 6. HELPERS
# ============================================================
eps = 1e-8

def code_to_attempt(code):
    try:
        code = int(code)
    except:
        return np.nan

    if code in [1, 11]:
        return 1
    elif code in [2, 12]:
        return 2
    elif code in [3, 13]:
        return 3
    elif code in [4, 14]:
        return 4
    return np.nan

def safe_corr(a, b):
    a = np.asarray(a, dtype=float)
    b = np.asarray(b, dtype=float)
    if np.std(a) < 1e-12 or np.std(b) < 1e-12:
        return 0.0
    return np.corrcoef(a, b)[0, 1]

def linear_slope(x, y):
    x = np.asarray(x, dtype=float)
    y = np.asarray(y, dtype=float)
    if np.std(y) < 1e-12:
        return 0.0
    return np.polyfit(x, y, 1)[0]

def compute_train_amnesia_features(codes_pw_subset):
    n_participants, n_words = codes_pw_subset.shape

    am1_train = np.zeros(n_participants, dtype=float)
    amall_train = np.zeros(n_participants, dtype=float)
    mean_attempt_train = np.zeros(n_participants, dtype=float)
    max_attempt_train = np.zeros(n_participants, dtype=float)
    fail_rate_train = np.zeros(n_participants, dtype=float)
    delayed_rate_train = np.zeros(n_participants, dtype=float)

    for p in range(n_participants):
        row = codes_pw_subset[p]

        count_first_success = np.sum(row == 1)
        am1_train[p] = (n_words - count_first_success) / n_words

        count_total_fail = np.sum(np.isin(row, [11, 12, 13, 14]))
        amall_train[p] = count_total_fail / n_words

        attempts = np.array([code_to_attempt(x) for x in row], dtype=float)
        if np.all(np.isnan(attempts)):
            mean_attempt_train[p] = 0.0
            max_attempt_train[p] = 0.0
        else:
            mean_attempt_train[p] = np.nanmean(attempts)
            max_attempt_train[p] = np.nanmax(attempts)

        fail_rate_train[p] = count_total_fail / n_words
        delayed_rate_train[p] = np.sum(np.isin(row, [2, 3, 4])) / n_words

    return (
        am1_train,
        amall_train,
        mean_attempt_train,
        max_attempt_train,
        fail_rate_train,
        delayed_rate_train
    )

# ============================================================
# 7. WORD SPLIT
# ============================================================
rng = np.random.RandomState(42)
word_idx = np.arange(30)
rng.shuffle(word_idx)

train_word_idx = word_idx[:24]
test_word_idx = word_idx[24:30]

print("\nWord split")
print("Train words:", len(train_word_idx))
print("Test words:", len(test_word_idx))

# ============================================================
# 8. TRAIN-WORD ATTEMPT FEATURES
# ============================================================
codes_train = codes_pw[:, train_word_idx]

(
    am1_train,
    amall_train,
    mean_attempt_train,
    max_attempt_train,
    fail_rate_train,
    delayed_rate_train
) = compute_train_amnesia_features(codes_train)

# ============================================================
# 9. TARGET
# ============================================================
y_pt_proxy = np.mean(T_raw[:, test_word_idx], axis=1)

# ============================================================
# 10. PARTICIPANT SPLIT FIRST
# ============================================================
participant_ids = np.arange(1, 107)

train_idx, test_idx = train_test_split(
    np.arange(106),
    test_size=20,
    random_state=42
)

# training-only reference average to reduce leakage
avg_word_time_train_ref = np.mean(T_raw[train_idx], axis=0)

# ============================================================
# 11. FEATURE ENGINEERING
# ============================================================
def build_participant_features(
    T_raw, age, habit,
    am1_train, amall_train,
    mean_attempt_train, max_attempt_train, fail_rate_train, delayed_rate_train,
    ortho, avg_word_time_ref, train_word_idx, codes_pw
):
    Ttr = T_raw[:, train_word_idx]
    ortho_tr = ortho[train_word_idx]
    avg_time_tr = avg_word_time_ref[train_word_idx]

    n = Ttr.shape[0]
    idx = np.arange(1, Ttr.shape[1] + 1)

    attempts_tr = np.array([
        [code_to_attempt(x) for x in row]
        for row in codes_pw[:, train_word_idx]
    ], dtype=float)
    attempts_tr = np.nan_to_num(attempts_tr, nan=0.0)

    mean_t = np.mean(Ttr, axis=1)
    std_t = np.std(Ttr, axis=1)
    median_t = np.median(Ttr, axis=1)
    min_t = np.min(Ttr, axis=1)
    max_t = np.max(Ttr, axis=1)
    range_t = max_t - min_t
    cv_t = std_t / (mean_t + eps)

    slow_ratio = np.mean(Ttr > median_t[:, None], axis=1)
    extreme_ratio = np.mean(Ttr > (mean_t[:, None] + 2 * std_t[:, None]), axis=1)
    percentile_gap = np.percentile(Ttr, 90, axis=1) - np.percentile(Ttr, 10, axis=1)

    first8 = np.mean(Ttr[:, :8], axis=1)
    middle8 = np.mean(Ttr[:, 8:16], axis=1)
    last8 = np.mean(Ttr[:, 16:24], axis=1)
    late_minus_early = last8 - first8
    slope_trial = np.array([linear_slope(idx, Ttr[i]) for i in range(n)])

    norm_t = Ttr / (avg_time_tr.reshape(1, -1) + eps)
    mean_norm = np.mean(norm_t, axis=1)
    std_norm = np.std(norm_t, axis=1)
    max_norm = np.max(norm_t, axis=1)

    resid_t = Ttr - avg_time_tr.reshape(1, -1)
    mean_resid = np.mean(resid_t, axis=1)
    std_resid = np.std(resid_t, axis=1)

    weighted_ortho_time = np.sum(Ttr * ortho_tr.reshape(1, -1), axis=1) / (np.sum(ortho_tr) + eps)
    corr_time_ortho = np.array([safe_corr(Ttr[i], ortho_tr) for i in range(n)])
    corr_time_avg = np.array([safe_corr(Ttr[i], avg_time_tr) for i in range(n)])
    slope_ortho = np.array([linear_slope(ortho_tr, Ttr[i]) for i in range(n)])

    q1 = np.quantile(ortho_tr, 1/3)
    q2 = np.quantile(ortho_tr, 2/3)

    low_mask = ortho_tr <= q1
    mid_mask = (ortho_tr > q1) & (ortho_tr <= q2)
    high_mask = ortho_tr > q2

    mean_low = np.mean(Ttr[:, low_mask], axis=1)
    mean_mid = np.mean(Ttr[:, mid_mask], axis=1)
    mean_high = np.mean(Ttr[:, high_mask], axis=1)
    high_minus_low = mean_high - mean_low
    high_over_low = mean_high / (mean_low + eps)

    time_per_habit = mean_t / (habit + eps)
    age_time_interaction = age * mean_t
    habit_time_interaction = habit * mean_t

    df_feat = pd.DataFrame({
        "age": age,
        "habit": habit,

        "am1_train": am1_train,
        "amall_train": amall_train,
        "mean_attempt_train": mean_attempt_train,
        "max_attempt_train": max_attempt_train,
        "fail_rate_train": fail_rate_train,
        "delayed_rate_train": delayed_rate_train,

        "mean_t": mean_t,
        "std_t": std_t,
        "median_t": median_t,
        "min_t": min_t,
        "max_t": max_t,
        "range_t": range_t,
        "cv_t": cv_t,

        "slow_ratio": slow_ratio,
        "extreme_ratio": extreme_ratio,
        "percentile_gap": percentile_gap,

        "first8_t": first8,
        "middle8_t": middle8,
        "last8_t": last8,
        "late_minus_early": late_minus_early,
        "slope_trial": slope_trial,

        "mean_norm": mean_norm,
        "std_norm": std_norm,
        "max_norm": max_norm,
        "mean_resid": mean_resid,
        "std_resid": std_resid,

        "weighted_ortho_time": weighted_ortho_time,
        "corr_time_ortho": corr_time_ortho,
        "corr_time_avg": corr_time_avg,
        "slope_ortho": slope_ortho,

        "mean_low_ortho_time": mean_low,
        "mean_mid_ortho_time": mean_mid,
        "mean_high_ortho_time": mean_high,
        "high_minus_low": high_minus_low,
        "high_over_low": high_over_low,

        "time_per_habit": time_per_habit,
        "age_time_interaction": age_time_interaction,
        "habit_time_interaction": habit_time_interaction
    })

    for j in range(Ttr.shape[1]):
        df_feat[f"raw_time_{j+1:02d}"] = Ttr[:, j]

    for j in range(norm_t.shape[1]):
        df_feat[f"norm_time_{j+1:02d}"] = norm_t[:, j]

    for j in range(attempts_tr.shape[1]):
        df_feat[f"attempt_{j+1:02d}"] = attempts_tr[:, j]

    return df_feat.fillna(0)

X_feat = build_participant_features(
    T_raw, age, habit,
    am1_train, amall_train,
    mean_attempt_train, max_attempt_train, fail_rate_train, delayed_rate_train,
    ortho, avg_word_time_train_ref, train_word_idx, codes_pw
)

# ============================================================
# 12. TRAIN / TEST MATRICES
# ============================================================
X_train_full = X_feat.iloc[train_idx].reset_index(drop=True)
X_test_full = X_feat.iloc[test_idx].reset_index(drop=True)

y_train = y_pt_proxy[train_idx]
y_test = y_pt_proxy[test_idx]

id_test = participant_ids[test_idx]

# ============================================================
# 13. TRAIN-ONLY FEATURE SELECTION
# ============================================================
def select_top_features(X_df, y, top_k=40):
    corrs = []
    for col in X_df.columns:
        c = safe_corr(X_df[col].values, y)
        corrs.append(abs(c))
    corrs = np.array(corrs)
    idx = np.argsort(corrs)[-top_k:]
    cols = list(X_df.columns[idx])
    return cols, pd.Series(corrs, index=X_df.columns).sort_values(ascending=False)

TOP_K = 40
selected_cols, corr_series = select_top_features(X_train_full, y_train, top_k=TOP_K)

X_train = X_train_full[selected_cols].copy()
X_test = X_test_full[selected_cols].copy()

print("\nSelected feature matrix shape (train):", X_train.shape)
print("Selected feature matrix shape (test):", X_test.shape)
print("\nTop correlated training features:")
print(corr_series.head(15))

# ============================================================
# 14. RANDOM FOREST REGRESSOR ONLY
# ============================================================
rfr_pipe = Pipeline([
    ("imputer", SimpleImputer(strategy="median")),
    ("model", RandomForestRegressor())
])

rfr_param_grid = {}


cv = KFold(n_splits=5, shuffle=True, random_state=42)

search = GridSearchCV(
    estimator=rfr_pipe,
    param_grid=rfr_param_grid,
    scoring="r2",
    cv=cv,
    n_jobs=-1,
    refit=True,
    error_score="raise"
)

print("\n==================== RFR ONLY ====================")
search.fit(X_train, y_train)

y_pred = search.best_estimator_.predict(X_test)

mae = mean_absolute_error(y_test, y_pred)
mse = mean_squared_error(y_test, y_pred)
rmse = np.sqrt(mse)
r2 = r2_score(y_test, y_pred)

print("\nBest Params:")
print(search.best_params_)

print("\nMetrics:")
print(f"BestCV_R2 = {search.best_score_:.4f}")
print(f"Test_MAE  = {mae:.4f}")
print(f"Test_MSE  = {mse:.4f}")
print(f"Test_RMSE = {rmse:.4f}")
print(f"Test_R2   = {r2:.4f}")

pred_df = pd.DataFrame({
    "participant_id": id_test,
    "true_avg_time_test6": y_test,
    "pred_avg_time_test6": y_pred
})

print("\nPredictions preview:")
print(pred_df.head())

# Optional save
# pred_df.to_csv("/content/rfr_only_predictions.csv", index=False)
# print("\nSaved: /content/rfr_only_predictions.csv")

# ============================================================
# GBR-ONLY WRITING TIME PREDICTION
# ============================================================

import warnings
warnings.filterwarnings("ignore")

import numpy as np
import pandas as pd
from openpyxl import load_workbook

from sklearn.model_selection import train_test_split, KFold, GridSearchCV
from sklearn.impute import SimpleImputer
from sklearn.pipeline import Pipeline
from sklearn.metrics import mean_absolute_error, mean_squared_error, r2_score
from sklearn.ensemble import GradientBoostingRegressor

# ============================================================
# 1. FILE SETTINGS
# ============================================================
FILE_PATH = "<insert file path here>"
SHEET1 = "Sheet1"
SHEET2 = "Sheet2"

# ============================================================
# 2. EXCEL LAYOUT
# ============================================================
S1_START_COL = 2
S1_END_COL = 107
S1_FEATURE_ROW_START = 3
S1_FEATURE_ROW_END = 34

S2_WORD_COL = 1
S2_ORTHO_COL = 2
S2_AVGTIME_COL = 3
S2_CODES_START_COL = 4
S2_CODES_END_COL = 109
S2_ROW_START = 1
S2_ROW_END = 60

# ============================================================
# 3. LOAD WORKBOOK
# ============================================================
wb = load_workbook(FILE_PATH, data_only=True)
ws1 = wb[SHEET1]
ws2 = wb[SHEET2]

# ============================================================
# 4. READ SHEET1
# ============================================================
sheet1_rows = []
for r in range(S1_FEATURE_ROW_START, S1_FEATURE_ROW_END + 1):
    vals = []
    for c in range(S1_START_COL, S1_END_COL + 1):
        vals.append(ws1.cell(row=r, column=c).value)
    sheet1_rows.append(vals)

sheet1_rows = np.array(sheet1_rows, dtype=float)   # (32, 106)
X_sheet1 = sheet1_rows.T                           # (106, 32)

T_raw = X_sheet1[:, :30]
age = X_sheet1[:, 30]
habit = X_sheet1[:, 31]

print("Sheet1 loaded")
print("X_sheet1 shape:", X_sheet1.shape)
print("T_raw shape:", T_raw.shape)
print("age shape:", age.shape)
print("habit shape:", habit.shape)

# ============================================================
# 5. READ SHEET2
# ============================================================
words = []
ortho = []
avg_word_time = []
code_rows = []

for r in range(S2_ROW_START, S2_ROW_END + 1):
    word = ws2.cell(row=r, column=S2_WORD_COL).value
    osyll = ws2.cell(row=r, column=S2_ORTHO_COL).value
    tsec = ws2.cell(row=r, column=S2_AVGTIME_COL).value

    if word is not None and osyll is not None and tsec is not None:
        words.append(str(word))
        ortho.append(float(osyll))
        avg_word_time.append(float(tsec))

        row_codes = []
        for c in range(S2_CODES_START_COL, S2_CODES_END_COL + 1):
            row_codes.append(ws2.cell(row=r, column=c).value)
        code_rows.append(row_codes)

words = words[:30]
ortho = np.array(ortho[:30], dtype=float)
avg_word_time = np.array(avg_word_time[:30], dtype=float)
code_rows = np.array(code_rows[:30], dtype=float)

assert len(words) == 30, f"Expected 30 words, got {len(words)}"
assert code_rows.shape == (30, 106), f"Expected (30,106), got {code_rows.shape}"

codes_pw = code_rows.T  # (106, 30)

print("\nSheet2 loaded")
print("Number of words:", len(words))
print("ortho shape:", ortho.shape)
print("avg_word_time shape:", avg_word_time.shape)
print("codes_pw shape:", codes_pw.shape)

# ============================================================
# 6. HELPERS
# ============================================================
eps = 1e-8

def code_to_attempt(code):
    try:
        code = int(code)
    except:
        return np.nan

    if code in [1, 11]:
        return 1
    elif code in [2, 12]:
        return 2
    elif code in [3, 13]:
        return 3
    elif code in [4, 14]:
        return 4
    return np.nan

def safe_corr(a, b):
    a = np.asarray(a, dtype=float)
    b = np.asarray(b, dtype=float)
    if np.std(a) < 1e-12 or np.std(b) < 1e-12:
        return 0.0
    return np.corrcoef(a, b)[0, 1]

def linear_slope(x, y):
    x = np.asarray(x, dtype=float)
    y = np.asarray(y, dtype=float)
    if np.std(y) < 1e-12:
        return 0.0
    return np.polyfit(x, y, 1)[0]

def compute_train_amnesia_features(codes_pw_subset):
    n_participants, n_words = codes_pw_subset.shape

    am1_train = np.zeros(n_participants, dtype=float)
    amall_train = np.zeros(n_participants, dtype=float)
    mean_attempt_train = np.zeros(n_participants, dtype=float)
    max_attempt_train = np.zeros(n_participants, dtype=float)
    fail_rate_train = np.zeros(n_participants, dtype=float)
    delayed_rate_train = np.zeros(n_participants, dtype=float)

    for p in range(n_participants):
        row = codes_pw_subset[p]

        count_first_success = np.sum(row == 1)
        am1_train[p] = (n_words - count_first_success) / n_words

        count_total_fail = np.sum(np.isin(row, [11, 12, 13, 14]))
        amall_train[p] = count_total_fail / n_words

        attempts = np.array([code_to_attempt(x) for x in row], dtype=float)
        if np.all(np.isnan(attempts)):
            mean_attempt_train[p] = 0.0
            max_attempt_train[p] = 0.0
        else:
            mean_attempt_train[p] = np.nanmean(attempts)
            max_attempt_train[p] = np.nanmax(attempts)

        fail_rate_train[p] = count_total_fail / n_words
        delayed_rate_train[p] = np.sum(np.isin(row, [2, 3, 4])) / n_words

    return (
        am1_train,
        amall_train,
        mean_attempt_train,
        max_attempt_train,
        fail_rate_train,
        delayed_rate_train
    )

# ============================================================
# 7. WORD SPLIT
# ============================================================
rng = np.random.RandomState(42)
word_idx = np.arange(30)
rng.shuffle(word_idx)

train_word_idx = word_idx[:24]
test_word_idx = word_idx[24:30]

print("\nWord split")
print("Train words:", len(train_word_idx))
print("Test words:", len(test_word_idx))

# ============================================================
# 8. TRAIN-WORD ATTEMPT FEATURES
# ============================================================
codes_train = codes_pw[:, train_word_idx]

(
    am1_train,
    amall_train,
    mean_attempt_train,
    max_attempt_train,
    fail_rate_train,
    delayed_rate_train
) = compute_train_amnesia_features(codes_train)

# ============================================================
# 9. TARGET
# ============================================================
y_pt_proxy = np.mean(T_raw[:, test_word_idx], axis=1)

# ============================================================
# 10. PARTICIPANT SPLIT FIRST
# ============================================================
participant_ids = np.arange(1, 107)

train_idx, test_idx = train_test_split(
    np.arange(106),
    test_size=20,
    random_state=42
)

# training-only reference average to reduce leakage
avg_word_time_train_ref = np.mean(T_raw[train_idx], axis=0)

# ============================================================
# 11. FEATURE ENGINEERING
# ============================================================
def build_participant_features(
    T_raw, age, habit,
    am1_train, amall_train,
    mean_attempt_train, max_attempt_train, fail_rate_train, delayed_rate_train,
    ortho, avg_word_time_ref, train_word_idx, codes_pw
):
    Ttr = T_raw[:, train_word_idx]
    ortho_tr = ortho[train_word_idx]
    avg_time_tr = avg_word_time_ref[train_word_idx]

    n = Ttr.shape[0]
    idx = np.arange(1, Ttr.shape[1] + 1)

    attempts_tr = np.array([
        [code_to_attempt(x) for x in row]
        for row in codes_pw[:, train_word_idx]
    ], dtype=float)
    attempts_tr = np.nan_to_num(attempts_tr, nan=0.0)

    mean_t = np.mean(Ttr, axis=1)
    std_t = np.std(Ttr, axis=1)
    median_t = np.median(Ttr, axis=1)
    min_t = np.min(Ttr, axis=1)
    max_t = np.max(Ttr, axis=1)
    range_t = max_t - min_t
    cv_t = std_t / (mean_t + eps)

    slow_ratio = np.mean(Ttr > median_t[:, None], axis=1)
    extreme_ratio = np.mean(Ttr > (mean_t[:, None] + 2 * std_t[:, None]), axis=1)
    percentile_gap = np.percentile(Ttr, 90, axis=1) - np.percentile(Ttr, 10, axis=1)

    first8 = np.mean(Ttr[:, :8], axis=1)
    middle8 = np.mean(Ttr[:, 8:16], axis=1)
    last8 = np.mean(Ttr[:, 16:24], axis=1)
    late_minus_early = last8 - first8
    slope_trial = np.array([linear_slope(idx, Ttr[i]) for i in range(n)])

    norm_t = Ttr / (avg_time_tr.reshape(1, -1) + eps)
    mean_norm = np.mean(norm_t, axis=1)
    std_norm = np.std(norm_t, axis=1)
    max_norm = np.max(norm_t, axis=1)

    resid_t = Ttr - avg_time_tr.reshape(1, -1)
    mean_resid = np.mean(resid_t, axis=1)
    std_resid = np.std(resid_t, axis=1)

    weighted_ortho_time = np.sum(Ttr * ortho_tr.reshape(1, -1), axis=1) / (np.sum(ortho_tr) + eps)
    corr_time_ortho = np.array([safe_corr(Ttr[i], ortho_tr) for i in range(n)])
    corr_time_avg = np.array([safe_corr(Ttr[i], avg_time_tr) for i in range(n)])
    slope_ortho = np.array([linear_slope(ortho_tr, Ttr[i]) for i in range(n)])

    q1 = np.quantile(ortho_tr, 1/3)
    q2 = np.quantile(ortho_tr, 2/3)

    low_mask = ortho_tr <= q1
    mid_mask = (ortho_tr > q1) & (ortho_tr <= q2)
    high_mask = ortho_tr > q2

    mean_low = np.mean(Ttr[:, low_mask], axis=1)
    mean_mid = np.mean(Ttr[:, mid_mask], axis=1)
    mean_high = np.mean(Ttr[:, high_mask], axis=1)
    high_minus_low = mean_high - mean_low
    high_over_low = mean_high / (mean_low + eps)

    time_per_habit = mean_t / (habit + eps)
    age_time_interaction = age * mean_t
    habit_time_interaction = habit * mean_t

    df_feat = pd.DataFrame({
        "age": age,
        "habit": habit,

        "am1_train": am1_train,
        "amall_train": amall_train,
        "mean_attempt_train": mean_attempt_train,
        "max_attempt_train": max_attempt_train,
        "fail_rate_train": fail_rate_train,
        "delayed_rate_train": delayed_rate_train,

        "mean_t": mean_t,
        "std_t": std_t,
        "median_t": median_t,
        "min_t": min_t,
        "max_t": max_t,
        "range_t": range_t,
        "cv_t": cv_t,

        "slow_ratio": slow_ratio,
        "extreme_ratio": extreme_ratio,
        "percentile_gap": percentile_gap,

        "first8_t": first8,
        "middle8_t": middle8,
        "last8_t": last8,
        "late_minus_early": late_minus_early,
        "slope_trial": slope_trial,

        "mean_norm": mean_norm,
        "std_norm": std_norm,
        "max_norm": max_norm,
        "mean_resid": mean_resid,
        "std_resid": std_resid,

        "weighted_ortho_time": weighted_ortho_time,
        "corr_time_ortho": corr_time_ortho,
        "corr_time_avg": corr_time_avg,
        "slope_ortho": slope_ortho,

        "mean_low_ortho_time": mean_low,
        "mean_mid_ortho_time": mean_mid,
        "mean_high_ortho_time": mean_high,
        "high_minus_low": high_minus_low,
        "high_over_low": high_over_low,

        "time_per_habit": time_per_habit,
        "age_time_interaction": age_time_interaction,
        "habit_time_interaction": habit_time_interaction
    })

    for j in range(Ttr.shape[1]):
        df_feat[f"raw_time_{j+1:02d}"] = Ttr[:, j]

    for j in range(norm_t.shape[1]):
        df_feat[f"norm_time_{j+1:02d}"] = norm_t[:, j]

    for j in range(attempts_tr.shape[1]):
        df_feat[f"attempt_{j+1:02d}"] = attempts_tr[:, j]

    return df_feat.fillna(0)

X_feat = build_participant_features(
    T_raw, age, habit,
    am1_train, amall_train,
    mean_attempt_train, max_attempt_train, fail_rate_train, delayed_rate_train,
    ortho, avg_word_time_train_ref, train_word_idx, codes_pw
)

# ============================================================
# 12. TRAIN / TEST MATRICES
# ============================================================
X_train_full = X_feat.iloc[train_idx].reset_index(drop=True)
X_test_full = X_feat.iloc[test_idx].reset_index(drop=True)

y_train = y_pt_proxy[train_idx]
y_test = y_pt_proxy[test_idx]

id_test = participant_ids[test_idx]

# ============================================================
# 13. TRAIN-ONLY FEATURE SELECTION
# ============================================================
def select_top_features(X_df, y, top_k=40):
    corrs = []
    for col in X_df.columns:
        c = safe_corr(X_df[col].values, y)
        corrs.append(abs(c))
    corrs = np.array(corrs)
    idx = np.argsort(corrs)[-top_k:]
    cols = list(X_df.columns[idx])
    return cols, pd.Series(corrs, index=X_df.columns).sort_values(ascending=False)

TOP_K = 40
selected_cols, corr_series = select_top_features(X_train_full, y_train, top_k=TOP_K)

X_train = X_train_full[selected_cols].copy()
X_test = X_test_full[selected_cols].copy()

print("\nSelected feature matrix shape (train):", X_train.shape)
print("Selected feature matrix shape (test):", X_test.shape)
print("\nTop correlated training features:")
print(corr_series.head(15))

# ============================================================
# 14. GBR ONLY
# ============================================================
gbr_pipe = Pipeline([
    ("imputer", SimpleImputer(strategy="median")),
    ("model", GradientBoostingRegressor(random_state=42))
])

gbr_param_grid = {}

cv = KFold(n_splits=5, shuffle=True, random_state=42)

search = GridSearchCV(
    estimator=gbr_pipe,
    param_grid=gbr_param_grid,
    scoring="r2",
    cv=cv,
    n_jobs=-1,
    refit=True,
    error_score="raise"
)

print("\n==================== GBR ONLY ====================")
search.fit(X_train, y_train)

y_pred = search.best_estimator_.predict(X_test)

mae = mean_absolute_error(y_test, y_pred)
mse = mean_squared_error(y_test, y_pred)
rmse = np.sqrt(mse)
r2 = r2_score(y_test, y_pred)

print("\nBest Params:")
print(search.best_params_)

print("\nMetrics:")
print(f"BestCV_R2 = {search.best_score_:.4f}")
print(f"Test_MAE  = {mae:.4f}")
print(f"Test_MSE  = {mse:.4f}")
print(f"Test_RMSE = {rmse:.4f}")
print(f"Test_R2   = {r2:.4f}")

pred_df = pd.DataFrame({
    "participant_id": id_test,
    "true_avg_time_test6": y_test,
    "pred_avg_time_test6": y_pred
})

print("\nPredictions preview:")
print(pred_df.head())

# Optional save
# pred_df.to_csv("/content/gbr_only_predictions.csv", index=False)
# print("\nSaved: /content/gbr_only_predictions.csv")

# ============================================================
# ADABOOST-ONLY WRITING TIME PREDICTION
# ============================================================

import warnings
warnings.filterwarnings("ignore")

import numpy as np
import pandas as pd
from openpyxl import load_workbook

from sklearn.model_selection import train_test_split, KFold, GridSearchCV
from sklearn.impute import SimpleImputer
from sklearn.pipeline import Pipeline
from sklearn.metrics import mean_absolute_error, mean_squared_error, r2_score
from sklearn.ensemble import AdaBoostRegressor

# ============================================================
# 1. FILE SETTINGS
# ============================================================
FILE_PATH = "<insert file path here>"
SHEET1 = "Sheet1"
SHEET2 = "Sheet2"

# ============================================================
# 2. EXCEL LAYOUT
# ============================================================
S1_START_COL = 2
S1_END_COL = 107
S1_FEATURE_ROW_START = 3
S1_FEATURE_ROW_END = 34

S2_WORD_COL = 1
S2_ORTHO_COL = 2
S2_AVGTIME_COL = 3
S2_CODES_START_COL = 4
S2_CODES_END_COL = 109
S2_ROW_START = 1
S2_ROW_END = 60

# ============================================================
# 3. LOAD WORKBOOK
# ============================================================
wb = load_workbook(FILE_PATH, data_only=True)
ws1 = wb[SHEET1]
ws2 = wb[SHEET2]

# ============================================================
# 4. READ SHEET1
# ============================================================
sheet1_rows = []
for r in range(S1_FEATURE_ROW_START, S1_FEATURE_ROW_END + 1):
    vals = []
    for c in range(S1_START_COL, S1_END_COL + 1):
        vals.append(ws1.cell(row=r, column=c).value)
    sheet1_rows.append(vals)

sheet1_rows = np.array(sheet1_rows, dtype=float)   # (32, 106)
X_sheet1 = sheet1_rows.T                           # (106, 32)

T_raw = X_sheet1[:, :30]
age = X_sheet1[:, 30]
habit = X_sheet1[:, 31]

print("Sheet1 loaded")
print("X_sheet1 shape:", X_sheet1.shape)
print("T_raw shape:", T_raw.shape)
print("age shape:", age.shape)
print("habit shape:", habit.shape)

# ============================================================
# 5. READ SHEET2
# ============================================================
words = []
ortho = []
avg_word_time = []
code_rows = []

for r in range(S2_ROW_START, S2_ROW_END + 1):
    word = ws2.cell(row=r, column=S2_WORD_COL).value
    osyll = ws2.cell(row=r, column=S2_ORTHO_COL).value
    tsec = ws2.cell(row=r, column=S2_AVGTIME_COL).value

    if word is not None and osyll is not None and tsec is not None:
        words.append(str(word))
        ortho.append(float(osyll))
        avg_word_time.append(float(tsec))

        row_codes = []
        for c in range(S2_CODES_START_COL, S2_CODES_END_COL + 1):
            row_codes.append(ws2.cell(row=r, column=c).value)
        code_rows.append(row_codes)

words = words[:30]
ortho = np.array(ortho[:30], dtype=float)
avg_word_time = np.array(avg_word_time[:30], dtype=float)
code_rows = np.array(code_rows[:30], dtype=float)

assert len(words) == 30, f"Expected 30 words, got {len(words)}"
assert code_rows.shape == (30, 106), f"Expected (30,106), got {code_rows.shape}"

codes_pw = code_rows.T  # (106, 30)

print("\nSheet2 loaded")
print("Number of words:", len(words))
print("ortho shape:", ortho.shape)
print("avg_word_time shape:", avg_word_time.shape)
print("codes_pw shape:", codes_pw.shape)

# ============================================================
# 6. HELPERS
# ============================================================
eps = 1e-8

def code_to_attempt(code):
    try:
        code = int(code)
    except:
        return np.nan

    if code in [1, 11]:
        return 1
    elif code in [2, 12]:
        return 2
    elif code in [3, 13]:
        return 3
    elif code in [4, 14]:
        return 4
    return np.nan

def safe_corr(a, b):
    a = np.asarray(a, dtype=float)
    b = np.asarray(b, dtype=float)
    if np.std(a) < 1e-12 or np.std(b) < 1e-12:
        return 0.0
    return np.corrcoef(a, b)[0, 1]

def linear_slope(x, y):
    x = np.asarray(x, dtype=float)
    y = np.asarray(y, dtype=float)
    if np.std(y) < 1e-12:
        return 0.0
    return np.polyfit(x, y, 1)[0]

def compute_train_amnesia_features(codes_pw_subset):
    n_participants, n_words = codes_pw_subset.shape

    am1_train = np.zeros(n_participants, dtype=float)
    amall_train = np.zeros(n_participants, dtype=float)
    mean_attempt_train = np.zeros(n_participants, dtype=float)
    max_attempt_train = np.zeros(n_participants, dtype=float)
    fail_rate_train = np.zeros(n_participants, dtype=float)
    delayed_rate_train = np.zeros(n_participants, dtype=float)

    for p in range(n_participants):
        row = codes_pw_subset[p]

        count_first_success = np.sum(row == 1)
        am1_train[p] = (n_words - count_first_success) / n_words

        count_total_fail = np.sum(np.isin(row, [11, 12, 13, 14]))
        amall_train[p] = count_total_fail / n_words

        attempts = np.array([code_to_attempt(x) for x in row], dtype=float)
        if np.all(np.isnan(attempts)):
            mean_attempt_train[p] = 0.0
            max_attempt_train[p] = 0.0
        else:
            mean_attempt_train[p] = np.nanmean(attempts)
            max_attempt_train[p] = np.nanmax(attempts)

        fail_rate_train[p] = count_total_fail / n_words
        delayed_rate_train[p] = np.sum(np.isin(row, [2, 3, 4])) / n_words

    return (
        am1_train,
        amall_train,
        mean_attempt_train,
        max_attempt_train,
        fail_rate_train,
        delayed_rate_train
    )

# ============================================================
# 7. WORD SPLIT
# ============================================================
rng = np.random.RandomState(42)
word_idx = np.arange(30)
rng.shuffle(word_idx)

train_word_idx = word_idx[:24]
test_word_idx = word_idx[24:30]

print("\nWord split")
print("Train words:", len(train_word_idx))
print("Test words:", len(test_word_idx))

# ============================================================
# 8. TRAIN-WORD ATTEMPT FEATURES
# ============================================================
codes_train = codes_pw[:, train_word_idx]

(
    am1_train,
    amall_train,
    mean_attempt_train,
    max_attempt_train,
    fail_rate_train,
    delayed_rate_train
) = compute_train_amnesia_features(codes_train)

# ============================================================
# 9. TARGET
# ============================================================
y_pt_proxy = np.mean(T_raw[:, test_word_idx], axis=1)

# ============================================================
# 10. PARTICIPANT SPLIT FIRST
# ============================================================
participant_ids = np.arange(1, 107)

train_idx, test_idx = train_test_split(
    np.arange(106),
    test_size=20,
    random_state=42
)

# training-only reference average to reduce leakage
avg_word_time_train_ref = np.mean(T_raw[train_idx], axis=0)

# ============================================================
# 11. FEATURE ENGINEERING
# ============================================================
def build_participant_features(
    T_raw, age, habit,
    am1_train, amall_train,
    mean_attempt_train, max_attempt_train, fail_rate_train, delayed_rate_train,
    ortho, avg_word_time_ref, train_word_idx, codes_pw
):
    Ttr = T_raw[:, train_word_idx]
    ortho_tr = ortho[train_word_idx]
    avg_time_tr = avg_word_time_ref[train_word_idx]

    n = Ttr.shape[0]
    idx = np.arange(1, Ttr.shape[1] + 1)

    attempts_tr = np.array([
        [code_to_attempt(x) for x in row]
        for row in codes_pw[:, train_word_idx]
    ], dtype=float)
    attempts_tr = np.nan_to_num(attempts_tr, nan=0.0)

    mean_t = np.mean(Ttr, axis=1)
    std_t = np.std(Ttr, axis=1)
    median_t = np.median(Ttr, axis=1)
    min_t = np.min(Ttr, axis=1)
    max_t = np.max(Ttr, axis=1)
    range_t = max_t - min_t
    cv_t = std_t / (mean_t + eps)

    slow_ratio = np.mean(Ttr > median_t[:, None], axis=1)
    extreme_ratio = np.mean(Ttr > (mean_t[:, None] + 2 * std_t[:, None]), axis=1)
    percentile_gap = np.percentile(Ttr, 90, axis=1) - np.percentile(Ttr, 10, axis=1)

    first8 = np.mean(Ttr[:, :8], axis=1)
    middle8 = np.mean(Ttr[:, 8:16], axis=1)
    last8 = np.mean(Ttr[:, 16:24], axis=1)
    late_minus_early = last8 - first8
    slope_trial = np.array([linear_slope(idx, Ttr[i]) for i in range(n)])

    norm_t = Ttr / (avg_time_tr.reshape(1, -1) + eps)
    mean_norm = np.mean(norm_t, axis=1)
    std_norm = np.std(norm_t, axis=1)
    max_norm = np.max(norm_t, axis=1)

    resid_t = Ttr - avg_time_tr.reshape(1, -1)
    mean_resid = np.mean(resid_t, axis=1)
    std_resid = np.std(resid_t, axis=1)

    weighted_ortho_time = np.sum(Ttr * ortho_tr.reshape(1, -1), axis=1) / (np.sum(ortho_tr) + eps)
    corr_time_ortho = np.array([safe_corr(Ttr[i], ortho_tr) for i in range(n)])
    corr_time_avg = np.array([safe_corr(Ttr[i], avg_time_tr) for i in range(n)])
    slope_ortho = np.array([linear_slope(ortho_tr, Ttr[i]) for i in range(n)])

    q1 = np.quantile(ortho_tr, 1/3)
    q2 = np.quantile(ortho_tr, 2/3)

    low_mask = ortho_tr <= q1
    mid_mask = (ortho_tr > q1) & (ortho_tr <= q2)
    high_mask = ortho_tr > q2

    mean_low = np.mean(Ttr[:, low_mask], axis=1)
    mean_mid = np.mean(Ttr[:, mid_mask], axis=1)
    mean_high = np.mean(Ttr[:, high_mask], axis=1)
    high_minus_low = mean_high - mean_low
    high_over_low = mean_high / (mean_low + eps)

    time_per_habit = mean_t / (habit + eps)
    age_time_interaction = age * mean_t
    habit_time_interaction = habit * mean_t

    df_feat = pd.DataFrame({
        "age": age,
        "habit": habit,

        "am1_train": am1_train,
        "amall_train": amall_train,
        "mean_attempt_train": mean_attempt_train,
        "max_attempt_train": max_attempt_train,
        "fail_rate_train": fail_rate_train,
        "delayed_rate_train": delayed_rate_train,

        "mean_t": mean_t,
        "std_t": std_t,
        "median_t": median_t,
        "min_t": min_t,
        "max_t": max_t,
        "range_t": range_t,
        "cv_t": cv_t,

        "slow_ratio": slow_ratio,
        "extreme_ratio": extreme_ratio,
        "percentile_gap": percentile_gap,

        "first8_t": first8,
        "middle8_t": middle8,
        "last8_t": last8,
        "late_minus_early": late_minus_early,
        "slope_trial": slope_trial,

        "mean_norm": mean_norm,
        "std_norm": std_norm,
        "max_norm": max_norm,
        "mean_resid": mean_resid,
        "std_resid": std_resid,

        "weighted_ortho_time": weighted_ortho_time,
        "corr_time_ortho": corr_time_ortho,
        "corr_time_avg": corr_time_avg,
        "slope_ortho": slope_ortho,

        "mean_low_ortho_time": mean_low,
        "mean_mid_ortho_time": mean_mid,
        "mean_high_ortho_time": mean_high,
        "high_minus_low": high_minus_low,
        "high_over_low": high_over_low,

        "time_per_habit": time_per_habit,
        "age_time_interaction": age_time_interaction,
        "habit_time_interaction": habit_time_interaction
    })

    for j in range(Ttr.shape[1]):
        df_feat[f"raw_time_{j+1:02d}"] = Ttr[:, j]

    for j in range(norm_t.shape[1]):
        df_feat[f"norm_time_{j+1:02d}"] = norm_t[:, j]

    for j in range(attempts_tr.shape[1]):
        df_feat[f"attempt_{j+1:02d}"] = attempts_tr[:, j]

    return df_feat.fillna(0)

X_feat = build_participant_features(
    T_raw, age, habit,
    am1_train, amall_train,
    mean_attempt_train, max_attempt_train, fail_rate_train, delayed_rate_train,
    ortho, avg_word_time_train_ref, train_word_idx, codes_pw
)

# ============================================================
# 12. TRAIN / TEST MATRICES
# ============================================================
X_train_full = X_feat.iloc[train_idx].reset_index(drop=True)
X_test_full = X_feat.iloc[test_idx].reset_index(drop=True)

y_train = y_pt_proxy[train_idx]
y_test = y_pt_proxy[test_idx]

id_test = participant_ids[test_idx]

# ============================================================
# 13. TRAIN-ONLY FEATURE SELECTION
# ============================================================
def select_top_features(X_df, y, top_k=40):
    corrs = []
    for col in X_df.columns:
        c = safe_corr(X_df[col].values, y)
        corrs.append(abs(c))
    corrs = np.array(corrs)
    idx = np.argsort(corrs)[-top_k:]
    cols = list(X_df.columns[idx])
    return cols, pd.Series(corrs, index=X_df.columns).sort_values(ascending=False)

TOP_K = 40
selected_cols, corr_series = select_top_features(X_train_full, y_train, top_k=TOP_K)

X_train = X_train_full[selected_cols].copy()
X_test = X_test_full[selected_cols].copy()

print("\nSelected feature matrix shape (train):", X_train.shape)
print("Selected feature matrix shape (test):", X_test.shape)
print("\nTop correlated training features:")
print(corr_series.head(15))

# ============================================================
# 14. ADABOOST ONLY
# ============================================================
ada_pipe = Pipeline([
    ("imputer", SimpleImputer(strategy="median")),
    ("model", AdaBoostRegressor(random_state=42))
])

ada_param_grid = {}

cv = KFold(n_splits=5, shuffle=True, random_state=42)

search = GridSearchCV(
    estimator=ada_pipe,
    param_grid=ada_param_grid,
    scoring="r2",
    cv=cv,
    n_jobs=-1,
    refit=True,
    error_score="raise"
)

print("\n==================== ADABOOST ONLY ====================")
search.fit(X_train, y_train)

y_pred = search.best_estimator_.predict(X_test)

mae = mean_absolute_error(y_test, y_pred)
mse = mean_squared_error(y_test, y_pred)
rmse = np.sqrt(mse)
r2 = r2_score(y_test, y_pred)

print("\nBest Params:")
print(search.best_params_)

print("\nMetrics:")
print(f"BestCV_R2 = {search.best_score_:.4f}")
print(f"Test_MAE  = {mae:.4f}")
print(f"Test_MSE  = {mse:.4f}")
print(f"Test_RMSE = {rmse:.4f}")
print(f"Test_R2   = {r2:.4f}")

pred_df = pd.DataFrame({
    "participant_id": id_test,
    "true_avg_time_test6": y_test,
    "pred_avg_time_test6": y_pred
})

print("\nPredictions preview:")
print(pred_df.head())

# Optional save
# pred_df.to_csv("/content/adaboost_only_predictions.csv", index=False)
# print("\nSaved: /content/adaboost_only_predictions.csv")

# ============================================================
# XGBOOST ONLY
# ============================================================
from sklearn.model_selection import KFold, RandomizedSearchCV
from sklearn.impute import SimpleImputer
from sklearn.pipeline import Pipeline
from sklearn.metrics import mean_absolute_error, mean_squared_error, r2_score
import xgboost as xgb
import numpy as np

# Keep the SAME feature matrix and SAME raw target that GBR used
# Example:
# X_train = ...
# X_test = ...
# y_train = y_pt_proxy[train_idx]
# y_test = y_pt_proxy[test_idx]

xgb_pipe = Pipeline([
    ("imputer", SimpleImputer(strategy="median")),
    ("model", xgb.XGBRegressor(
        random_state=42,
        objective="reg:squarederror",
        verbosity=0,
        tree_method="hist",
        n_jobs=-1
    ))
])

xgb_param_dist = {
    "model__n_estimators": [400, 600, 800, 1000],
    "model__learning_rate": [0.01, 0.02, 0.03],
    "model__max_depth": [2, 3],
    "model__min_child_weight": [1, 2],
    "model__subsample": [0.85, 1.0],
    "model__colsample_bytree": [0.85, 1.0],
    "model__gamma": [0.0],
    "model__reg_alpha": [0.0],
    "model__reg_lambda": [1.0]
}

cv = KFold(n_splits=5, shuffle=True, random_state=42)

search = RandomizedSearchCV(
    estimator=xgb_pipe,
    param_distributions=xgb_param_dist,
    n_iter=24,
    scoring="r2",
    cv=cv,
    n_jobs=-1,
    refit=True,
    random_state=42,
    error_score="raise"
)

search.fit(X_train, y_train)
y_pred = search.best_estimator_.predict(X_test)

mae = mean_absolute_error(y_test, y_pred)
mse = mean_squared_error(y_test, y_pred)
rmse = np.sqrt(mse)
r2 = r2_score(y_test, y_pred)

print("Best Params:")
print(search.best_params_)

print("\nMetrics:")
print(f"BestCV_R2 = {search.best_score_:.4f}")
print(f"Test_MAE  = {mae:.4f}")
print(f"Test_MSE  = {mse:.4f}")
print(f"Test_RMSE = {rmse:.4f}")
print(f"Test_R2   = {r2:.4f}")

# ============================================================
# LIGHTGBM-ONLY WRITING TIME PREDICTION

# ============================================================

import warnings
warnings.filterwarnings("ignore")

import numpy as np
import pandas as pd
from openpyxl import load_workbook

from sklearn.model_selection import train_test_split, KFold, GridSearchCV
from sklearn.impute import SimpleImputer
from sklearn.pipeline import Pipeline
from sklearn.metrics import mean_absolute_error, mean_squared_error, r2_score

from lightgbm import LGBMRegressor

# ============================================================
# 1. FILE SETTINGS
# ============================================================
FILE_PATH ="<insert file path here>"
SHEET1 = "Sheet1"
SHEET2 = "Sheet2"

# ============================================================
# 2. EXCEL LAYOUT
# ============================================================
S1_START_COL = 2
S1_END_COL = 107
S1_FEATURE_ROW_START = 3
S1_FEATURE_ROW_END = 34

S2_WORD_COL = 1
S2_ORTHO_COL = 2
S2_AVGTIME_COL = 3
S2_CODES_START_COL = 4
S2_CODES_END_COL = 109
S2_ROW_START = 1
S2_ROW_END = 60

# ============================================================
# 3. LOAD WORKBOOK
# ============================================================
wb = load_workbook(FILE_PATH, data_only=True)
ws1 = wb[SHEET1]
ws2 = wb[SHEET2]

# ============================================================
# 4. READ SHEET1
# ============================================================
sheet1_rows = []
for r in range(S1_FEATURE_ROW_START, S1_FEATURE_ROW_END + 1):
    vals = []
    for c in range(S1_START_COL, S1_END_COL + 1):
        vals.append(ws1.cell(row=r, column=c).value)
    sheet1_rows.append(vals)

sheet1_rows = np.array(sheet1_rows, dtype=float)   # (32, 106)
X_sheet1 = sheet1_rows.T                           # (106, 32)

T_raw = X_sheet1[:, :30]
age = X_sheet1[:, 30]
habit = X_sheet1[:, 31]

print("Sheet1 loaded")
print("X_sheet1 shape:", X_sheet1.shape)
print("T_raw shape:", T_raw.shape)
print("age shape:", age.shape)
print("habit shape:", habit.shape)

# ============================================================
# 5. READ SHEET2
# ============================================================
words = []
ortho = []
avg_word_time = []
code_rows = []

for r in range(S2_ROW_START, S2_ROW_END + 1):
    word = ws2.cell(row=r, column=S2_WORD_COL).value
    osyll = ws2.cell(row=r, column=S2_ORTHO_COL).value
    tsec = ws2.cell(row=r, column=S2_AVGTIME_COL).value

    if word is not None and osyll is not None and tsec is not None:
        words.append(str(word))
        ortho.append(float(osyll))
        avg_word_time.append(float(tsec))

        row_codes = []
        for c in range(S2_CODES_START_COL, S2_CODES_END_COL + 1):
            row_codes.append(ws2.cell(row=r, column=c).value)
        code_rows.append(row_codes)

words = words[:30]
ortho = np.array(ortho[:30], dtype=float)
avg_word_time = np.array(avg_word_time[:30], dtype=float)
code_rows = np.array(code_rows[:30], dtype=float)

assert len(words) == 30, f"Expected 30 words, got {len(words)}"
assert code_rows.shape == (30, 106), f"Expected (30,106), got {code_rows.shape}"

codes_pw = code_rows.T  # (106, 30)

print("\nSheet2 loaded")
print("Number of words:", len(words))
print("ortho shape:", ortho.shape)
print("avg_word_time shape:", avg_word_time.shape)
print("codes_pw shape:", codes_pw.shape)

# ============================================================
# 6. HELPERS
# ============================================================
eps = 1e-8

def code_to_attempt(code):
    try:
        code = int(code)
    except:
        return np.nan

    if code in [1, 11]:
        return 1
    elif code in [2, 12]:
        return 2
    elif code in [3, 13]:
        return 3
    elif code in [4, 14]:
        return 4
    return np.nan

def safe_corr(a, b):
    a = np.asarray(a, dtype=float)
    b = np.asarray(b, dtype=float)
    if np.std(a) < 1e-12 or np.std(b) < 1e-12:
        return 0.0
    return np.corrcoef(a, b)[0, 1]

def linear_slope(x, y):
    x = np.asarray(x, dtype=float)
    y = np.asarray(y, dtype=float)
    if np.std(y) < 1e-12:
        return 0.0
    return np.polyfit(x, y, 1)[0]

def compute_train_amnesia_features(codes_pw_subset):
    n_participants, n_words = codes_pw_subset.shape

    am1_train = np.zeros(n_participants, dtype=float)
    amall_train = np.zeros(n_participants, dtype=float)
    mean_attempt_train = np.zeros(n_participants, dtype=float)
    max_attempt_train = np.zeros(n_participants, dtype=float)
    fail_rate_train = np.zeros(n_participants, dtype=float)
    delayed_rate_train = np.zeros(n_participants, dtype=float)

    for p in range(n_participants):
        row = codes_pw_subset[p]

        count_first_success = np.sum(row == 1)
        am1_train[p] = (n_words - count_first_success) / n_words

        count_total_fail = np.sum(np.isin(row, [11, 12, 13, 14]))
        amall_train[p] = count_total_fail / n_words

        attempts = np.array([code_to_attempt(x) for x in row], dtype=float)
        if np.all(np.isnan(attempts)):
            mean_attempt_train[p] = 0.0
            max_attempt_train[p] = 0.0
        else:
            mean_attempt_train[p] = np.nanmean(attempts)
            max_attempt_train[p] = np.nanmax(attempts)

        fail_rate_train[p] = count_total_fail / n_words
        delayed_rate_train[p] = np.sum(np.isin(row, [2, 3, 4])) / n_words

    return (
        am1_train,
        amall_train,
        mean_attempt_train,
        max_attempt_train,
        fail_rate_train,
        delayed_rate_train
    )

# ============================================================
# 7. WORD SPLIT
# ============================================================
rng = np.random.RandomState(42)
word_idx = np.arange(30)
rng.shuffle(word_idx)

train_word_idx = word_idx[:24]
test_word_idx = word_idx[24:30]

print("\nWord split")
print("Train words:", len(train_word_idx))
print("Test words:", len(test_word_idx))

# ============================================================
# 8. TRAIN-WORD ATTEMPT FEATURES
# ============================================================
codes_train = codes_pw[:, train_word_idx]

(
    am1_train,
    amall_train,
    mean_attempt_train,
    max_attempt_train,
    fail_rate_train,
    delayed_rate_train
) = compute_train_amnesia_features(codes_train)

# ============================================================
# 9. TARGET
# ============================================================
y_pt_proxy = np.mean(T_raw[:, test_word_idx], axis=1)

# ============================================================
# 10. PARTICIPANT SPLIT FIRST
# ============================================================
participant_ids = np.arange(1, 107)

train_idx, test_idx = train_test_split(
    np.arange(106),
    test_size=20,
    random_state=42
)

# training-only reference average to reduce leakage
avg_word_time_train_ref = np.mean(T_raw[train_idx], axis=0)

# ============================================================
# 11. FEATURE ENGINEERING
# ============================================================
def build_participant_features(
    T_raw, age, habit,
    am1_train, amall_train,
    mean_attempt_train, max_attempt_train, fail_rate_train, delayed_rate_train,
    ortho, avg_word_time_ref, train_word_idx, codes_pw
):
    Ttr = T_raw[:, train_word_idx]
    ortho_tr = ortho[train_word_idx]
    avg_time_tr = avg_word_time_ref[train_word_idx]

    n = Ttr.shape[0]
    idx = np.arange(1, Ttr.shape[1] + 1)

    attempts_tr = np.array([
        [code_to_attempt(x) for x in row]
        for row in codes_pw[:, train_word_idx]
    ], dtype=float)
    attempts_tr = np.nan_to_num(attempts_tr, nan=0.0)

    mean_t = np.mean(Ttr, axis=1)
    std_t = np.std(Ttr, axis=1)
    median_t = np.median(Ttr, axis=1)
    min_t = np.min(Ttr, axis=1)
    max_t = np.max(Ttr, axis=1)
    range_t = max_t - min_t
    cv_t = std_t / (mean_t + eps)

    slow_ratio = np.mean(Ttr > median_t[:, None], axis=1)
    extreme_ratio = np.mean(Ttr > (mean_t[:, None] + 2 * std_t[:, None]), axis=1)
    percentile_gap = np.percentile(Ttr, 90, axis=1) - np.percentile(Ttr, 10, axis=1)

    first8 = np.mean(Ttr[:, :8], axis=1)
    middle8 = np.mean(Ttr[:, 8:16], axis=1)
    last8 = np.mean(Ttr[:, 16:24], axis=1)
    late_minus_early = last8 - first8
    slope_trial = np.array([linear_slope(idx, Ttr[i]) for i in range(n)])

    norm_t = Ttr / (avg_time_tr.reshape(1, -1) + eps)
    mean_norm = np.mean(norm_t, axis=1)
    std_norm = np.std(norm_t, axis=1)
    max_norm = np.max(norm_t, axis=1)

    resid_t = Ttr - avg_time_tr.reshape(1, -1)
    mean_resid = np.mean(resid_t, axis=1)
    std_resid = np.std(resid_t, axis=1)

    weighted_ortho_time = np.sum(Ttr * ortho_tr.reshape(1, -1), axis=1) / (np.sum(ortho_tr) + eps)
    corr_time_ortho = np.array([safe_corr(Ttr[i], ortho_tr) for i in range(n)])
    corr_time_avg = np.array([safe_corr(Ttr[i], avg_time_tr) for i in range(n)])
    slope_ortho = np.array([linear_slope(ortho_tr, Ttr[i]) for i in range(n)])

    q1 = np.quantile(ortho_tr, 1/3)
    q2 = np.quantile(ortho_tr, 2/3)

    low_mask = ortho_tr <= q1
    mid_mask = (ortho_tr > q1) & (ortho_tr <= q2)
    high_mask = ortho_tr > q2

    mean_low = np.mean(Ttr[:, low_mask], axis=1)
    mean_mid = np.mean(Ttr[:, mid_mask], axis=1)
    mean_high = np.mean(Ttr[:, high_mask], axis=1)
    high_minus_low = mean_high - mean_low
    high_over_low = mean_high / (mean_low + eps)

    time_per_habit = mean_t / (habit + eps)
    age_time_interaction = age * mean_t
    habit_time_interaction = habit * mean_t

    df_feat = pd.DataFrame({
        "age": age,
        "habit": habit,

        "am1_train": am1_train,
        "amall_train": amall_train,
        "mean_attempt_train": mean_attempt_train,
        "max_attempt_train": max_attempt_train,
        "fail_rate_train": fail_rate_train,
        "delayed_rate_train": delayed_rate_train,

        "mean_t": mean_t,
        "std_t": std_t,
        "median_t": median_t,
        "min_t": min_t,
        "max_t": max_t,
        "range_t": range_t,
        "cv_t": cv_t,

        "slow_ratio": slow_ratio,
        "extreme_ratio": extreme_ratio,
        "percentile_gap": percentile_gap,

        "first8_t": first8,
        "middle8_t": middle8,
        "last8_t": last8,
        "late_minus_early": late_minus_early,
        "slope_trial": slope_trial,

        "mean_norm": mean_norm,
        "std_norm": std_norm,
        "max_norm": max_norm,
        "mean_resid": mean_resid,
        "std_resid": std_resid,

        "weighted_ortho_time": weighted_ortho_time,
        "corr_time_ortho": corr_time_ortho,
        "corr_time_avg": corr_time_avg,
        "slope_ortho": slope_ortho,

        "mean_low_ortho_time": mean_low,
        "mean_mid_ortho_time": mean_mid,
        "mean_high_ortho_time": mean_high,
        "high_minus_low": high_minus_low,
        "high_over_low": high_over_low,

        "time_per_habit": time_per_habit,
        "age_time_interaction": age_time_interaction,
        "habit_time_interaction": habit_time_interaction
    })

    for j in range(Ttr.shape[1]):
        df_feat[f"raw_time_{j+1:02d}"] = Ttr[:, j]

    for j in range(norm_t.shape[1]):
        df_feat[f"norm_time_{j+1:02d}"] = norm_t[:, j]

    for j in range(attempts_tr.shape[1]):
        df_feat[f"attempt_{j+1:02d}"] = attempts_tr[:, j]

    return df_feat.fillna(0)

X_feat = build_participant_features(
    T_raw, age, habit,
    am1_train, amall_train,
    mean_attempt_train, max_attempt_train, fail_rate_train, delayed_rate_train,
    ortho, avg_word_time_train_ref, train_word_idx, codes_pw
)

# ============================================================
# 12. TRAIN / TEST MATRICES
# ============================================================
X_train_full = X_feat.iloc[train_idx].reset_index(drop=True)
X_test_full = X_feat.iloc[test_idx].reset_index(drop=True)

y_train = y_pt_proxy[train_idx]
y_test = y_pt_proxy[test_idx]

id_test = participant_ids[test_idx]

# ============================================================
# 13. TRAIN-ONLY FEATURE SELECTION
# ============================================================
def select_top_features(X_df, y, top_k=30):
    corrs = []
    for col in X_df.columns:
        c = safe_corr(X_df[col].values, y)
        corrs.append(abs(c))
    corrs = np.array(corrs)
    idx = np.argsort(corrs)[-top_k:]
    cols = list(X_df.columns[idx])
    return cols, pd.Series(corrs, index=X_df.columns).sort_values(ascending=False)

TOP_K = 30
selected_cols, corr_series = select_top_features(X_train_full, y_train, top_k=TOP_K)

X_train = X_train_full[selected_cols].copy()
X_test = X_test_full[selected_cols].copy()

print("\nSelected feature matrix shape (train):", X_train.shape)
print("Selected feature matrix shape (test):", X_test.shape)
print("\nTop correlated training features:")
print(corr_series.head(15))

# ============================================================
# 14. LIGHTGBM ONLY
# ============================================================
lgbm_pipe = Pipeline([
    ("imputer", SimpleImputer(strategy="median")),
    ("model", LGBMRegressor(
        random_state=42,
        verbose=-1,
        n_jobs=-1
    ))
])

lgbm_param_grid = {}

cv = KFold(n_splits=5, shuffle=True, random_state=42)

search = GridSearchCV(
    estimator=lgbm_pipe,
    param_grid=lgbm_param_grid,
    scoring="r2",
    cv=cv,
    n_jobs=-1,
    refit=True,
    error_score="raise"
)

print("\n==================== LIGHTGBM ONLY ====================")
search.fit(X_train, y_train)

y_pred = search.best_estimator_.predict(X_test)

mae = mean_absolute_error(y_test, y_pred)
mse = mean_squared_error(y_test, y_pred)
rmse = np.sqrt(mse)
r2 = r2_score(y_test, y_pred)

print("\nBest Params:")
print(search.best_params_)

print("\nMetrics:")
print(f"BestCV_R2 = {search.best_score_:.4f}")
print(f"Test_MAE  = {mae:.4f}")
print(f"Test_MSE  = {mse:.4f}")
print(f"Test_RMSE = {rmse:.4f}")
print(f"Test_R2   = {r2:.4f}")

pred_df = pd.DataFrame({
    "participant_id": id_test,
    "true_avg_time_test6": y_test,
    "pred_avg_time_test6": y_pred
})

print("\nPredictions preview:")
print(pred_df.head())

# Optional save
# pred_df.to_csv("/content/lightgbm_only_predictions.csv", index=False)
# print("\nSaved: /content/lightgbm_only_predictions.csv")

# ============================================================
# VOTING REGRESSOR-ONLY WRITING TIME PREDICTION
# richer time features + train-only feature selection + leakage control
# ============================================================

import warnings
warnings.filterwarnings("ignore")

import numpy as np
import pandas as pd
from openpyxl import load_workbook

from sklearn.model_selection import train_test_split, KFold, GridSearchCV
from sklearn.impute import SimpleImputer
from sklearn.pipeline import Pipeline
from sklearn.metrics import mean_absolute_error, mean_squared_error, r2_score
from sklearn.ensemble import (
    GradientBoostingRegressor,
    RandomForestRegressor,
    VotingRegressor
)

import xgboost as xgb

# ============================================================
# 1. FILE SETTINGS
# ============================================================
FILE_PATH = "<insert file path here>"
SHEET1 = "Sheet1"
SHEET2 = "Sheet2"

# ============================================================
# 2. EXCEL LAYOUT
# ============================================================
S1_START_COL = 2
S1_END_COL = 107
S1_FEATURE_ROW_START = 3
S1_FEATURE_ROW_END = 34

S2_WORD_COL = 1
S2_ORTHO_COL = 2
S2_AVGTIME_COL = 3
S2_CODES_START_COL = 4
S2_CODES_END_COL = 109
S2_ROW_START = 1
S2_ROW_END = 60

# ============================================================
# 3. LOAD WORKBOOK
# ============================================================
wb = load_workbook(FILE_PATH, data_only=True)
ws1 = wb[SHEET1]
ws2 = wb[SHEET2]

# ============================================================
# 4. READ SHEET1
# ============================================================
sheet1_rows = []
for r in range(S1_FEATURE_ROW_START, S1_FEATURE_ROW_END + 1):
    vals = []
    for c in range(S1_START_COL, S1_END_COL + 1):
        vals.append(ws1.cell(row=r, column=c).value)
    sheet1_rows.append(vals)

sheet1_rows = np.array(sheet1_rows, dtype=float)   # (32, 106)
X_sheet1 = sheet1_rows.T                           # (106, 32)

T_raw = X_sheet1[:, :30]
age = X_sheet1[:, 30]
habit = X_sheet1[:, 31]

print("Sheet1 loaded")
print("X_sheet1 shape:", X_sheet1.shape)
print("T_raw shape:", T_raw.shape)
print("age shape:", age.shape)
print("habit shape:", habit.shape)

# ============================================================
# 5. READ SHEET2
# ============================================================
words = []
ortho = []
avg_word_time = []
code_rows = []

for r in range(S2_ROW_START, S2_ROW_END + 1):
    word = ws2.cell(row=r, column=S2_WORD_COL).value
    osyll = ws2.cell(row=r, column=S2_ORTHO_COL).value
    tsec = ws2.cell(row=r, column=S2_AVGTIME_COL).value

    if word is not None and osyll is not None and tsec is not None:
        words.append(str(word))
        ortho.append(float(osyll))
        avg_word_time.append(float(tsec))

        row_codes = []
        for c in range(S2_CODES_START_COL, S2_CODES_END_COL + 1):
            row_codes.append(ws2.cell(row=r, column=c).value)
        code_rows.append(row_codes)

words = words[:30]
ortho = np.array(ortho[:30], dtype=float)
avg_word_time = np.array(avg_word_time[:30], dtype=float)
code_rows = np.array(code_rows[:30], dtype=float)

assert len(words) == 30, f"Expected 30 words, got {len(words)}"
assert code_rows.shape == (30, 106), f"Expected (30,106), got {code_rows.shape}"

codes_pw = code_rows.T

print("\nSheet2 loaded")
print("Number of words:", len(words))
print("ortho shape:", ortho.shape)
print("avg_word_time shape:", avg_word_time.shape)
print("codes_pw shape:", codes_pw.shape)

# ============================================================
# 6. HELPERS
# ============================================================
eps = 1e-8

def code_to_attempt(code):
    try:
        code = int(code)
    except:
        return np.nan

    if code in [1, 11]:
        return 1
    elif code in [2, 12]:
        return 2
    elif code in [3, 13]:
        return 3
    elif code in [4, 14]:
        return 4
    return np.nan

def safe_corr(a, b):
    a = np.asarray(a, dtype=float)
    b = np.asarray(b, dtype=float)
    if np.std(a) < 1e-12 or np.std(b) < 1e-12:
        return 0.0
    return np.corrcoef(a, b)[0, 1]

def linear_slope(x, y):
    x = np.asarray(x, dtype=float)
    y = np.asarray(y, dtype=float)
    if np.std(y) < 1e-12:
        return 0.0
    return np.polyfit(x, y, 1)[0]

def compute_train_amnesia_features(codes_pw_subset):
    n_participants, n_words = codes_pw_subset.shape

    am1_train = np.zeros(n_participants, dtype=float)
    amall_train = np.zeros(n_participants, dtype=float)
    mean_attempt_train = np.zeros(n_participants, dtype=float)
    max_attempt_train = np.zeros(n_participants, dtype=float)
    fail_rate_train = np.zeros(n_participants, dtype=float)
    delayed_rate_train = np.zeros(n_participants, dtype=float)

    for p in range(n_participants):
        row = codes_pw_subset[p]

        count_first_success = np.sum(row == 1)
        am1_train[p] = (n_words - count_first_success) / n_words

        count_total_fail = np.sum(np.isin(row, [11, 12, 13, 14]))
        amall_train[p] = count_total_fail / n_words

        attempts = np.array([code_to_attempt(x) for x in row], dtype=float)
        if np.all(np.isnan(attempts)):
            mean_attempt_train[p] = 0.0
            max_attempt_train[p] = 0.0
        else:
            mean_attempt_train[p] = np.nanmean(attempts)
            max_attempt_train[p] = np.nanmax(attempts)

        fail_rate_train[p] = count_total_fail / n_words
        delayed_rate_train[p] = np.sum(np.isin(row, [2, 3, 4])) / n_words

    return (
        am1_train,
        amall_train,
        mean_attempt_train,
        max_attempt_train,
        fail_rate_train,
        delayed_rate_train
    )

# ============================================================
# 7. WORD SPLIT
# ============================================================
rng = np.random.RandomState(42)
word_idx = np.arange(30)
rng.shuffle(word_idx)

train_word_idx = word_idx[:24]
test_word_idx = word_idx[24:30]

print("\nWord split")
print("Train words:", len(train_word_idx))
print("Test words:", len(test_word_idx))

# ============================================================
# 8. TRAIN-WORD ATTEMPT FEATURES
# ============================================================
codes_train = codes_pw[:, train_word_idx]

(
    am1_train,
    amall_train,
    mean_attempt_train,
    max_attempt_train,
    fail_rate_train,
    delayed_rate_train
) = compute_train_amnesia_features(codes_train)

# ============================================================
# 9. TARGET
# ============================================================
y_pt_proxy = np.mean(T_raw[:, test_word_idx], axis=1)

# ============================================================
# 10. PARTICIPANT SPLIT FIRST
# ============================================================
participant_ids = np.arange(1, 107)

train_idx, test_idx = train_test_split(
    np.arange(106),
    test_size=20,
    random_state=42
)

# training-only reference average to reduce leakage
avg_word_time_train_ref = np.mean(T_raw[train_idx], axis=0)

# ============================================================
# 11. FEATURE ENGINEERING
# ============================================================
def build_participant_features(
    T_raw, age, habit,
    am1_train, amall_train,
    mean_attempt_train, max_attempt_train, fail_rate_train, delayed_rate_train,
    ortho, avg_word_time_ref, train_word_idx, codes_pw
):
    Ttr = T_raw[:, train_word_idx]
    ortho_tr = ortho[train_word_idx]
    avg_time_tr = avg_word_time_ref[train_word_idx]

    n = Ttr.shape[0]
    idx = np.arange(1, Ttr.shape[1] + 1)

    attempts_tr = np.array([
        [code_to_attempt(x) for x in row]
        for row in codes_pw[:, train_word_idx]
    ], dtype=float)
    attempts_tr = np.nan_to_num(attempts_tr, nan=0.0)

    mean_t = np.mean(Ttr, axis=1)
    std_t = np.std(Ttr, axis=1)
    median_t = np.median(Ttr, axis=1)
    min_t = np.min(Ttr, axis=1)
    max_t = np.max(Ttr, axis=1)
    range_t = max_t - min_t
    cv_t = std_t / (mean_t + eps)

    slow_ratio = np.mean(Ttr > median_t[:, None], axis=1)
    extreme_ratio = np.mean(Ttr > (mean_t[:, None] + 2 * std_t[:, None]), axis=1)
    percentile_gap = np.percentile(Ttr, 90, axis=1) - np.percentile(Ttr, 10, axis=1)

    first8 = np.mean(Ttr[:, :8], axis=1)
    middle8 = np.mean(Ttr[:, 8:16], axis=1)
    last8 = np.mean(Ttr[:, 16:24], axis=1)
    late_minus_early = last8 - first8
    slope_trial = np.array([linear_slope(idx, Ttr[i]) for i in range(n)])

    norm_t = Ttr / (avg_time_tr.reshape(1, -1) + eps)
    mean_norm = np.mean(norm_t, axis=1)
    std_norm = np.std(norm_t, axis=1)
    max_norm = np.max(norm_t, axis=1)

    resid_t = Ttr - avg_time_tr.reshape(1, -1)
    mean_resid = np.mean(resid_t, axis=1)
    std_resid = np.std(resid_t, axis=1)

    weighted_ortho_time = np.sum(Ttr * ortho_tr.reshape(1, -1), axis=1) / (np.sum(ortho_tr) + eps)
    corr_time_ortho = np.array([safe_corr(Ttr[i], ortho_tr) for i in range(n)])
    corr_time_avg = np.array([safe_corr(Ttr[i], avg_time_tr) for i in range(n)])
    slope_ortho = np.array([linear_slope(ortho_tr, Ttr[i]) for i in range(n)])

    q1 = np.quantile(ortho_tr, 1/3)
    q2 = np.quantile(ortho_tr, 2/3)

    low_mask = ortho_tr <= q1
    mid_mask = (ortho_tr > q1) & (ortho_tr <= q2)
    high_mask = ortho_tr > q2

    mean_low = np.mean(Ttr[:, low_mask], axis=1)
    mean_mid = np.mean(Ttr[:, mid_mask], axis=1)
    mean_high = np.mean(Ttr[:, high_mask], axis=1)
    high_minus_low = mean_high - mean_low
    high_over_low = mean_high / (mean_low + eps)

    time_per_habit = mean_t / (habit + eps)
    age_time_interaction = age * mean_t
    habit_time_interaction = habit * mean_t

    df_feat = pd.DataFrame({
        "age": age,
        "habit": habit,

        "am1_train": am1_train,
        "amall_train": amall_train,
        "mean_attempt_train": mean_attempt_train,
        "max_attempt_train": max_attempt_train,
        "fail_rate_train": fail_rate_train,
        "delayed_rate_train": delayed_rate_train,

        "mean_t": mean_t,
        "std_t": std_t,
        "median_t": median_t,
        "min_t": min_t,
        "max_t": max_t,
        "range_t": range_t,
        "cv_t": cv_t,

        "slow_ratio": slow_ratio,
        "extreme_ratio": extreme_ratio,
        "percentile_gap": percentile_gap,

        "first8_t": first8,
        "middle8_t": middle8,
        "last8_t": last8,
        "late_minus_early": late_minus_early,
        "slope_trial": slope_trial,

        "mean_norm": mean_norm,
        "std_norm": std_norm,
        "max_norm": max_norm,
        "mean_resid": mean_resid,
        "std_resid": std_resid,

        "weighted_ortho_time": weighted_ortho_time,
        "corr_time_ortho": corr_time_ortho,
        "corr_time_avg": corr_time_avg,
        "slope_ortho": slope_ortho,

        "mean_low_ortho_time": mean_low,
        "mean_mid_ortho_time": mean_mid,
        "mean_high_ortho_time": mean_high,
        "high_minus_low": high_minus_low,
        "high_over_low": high_over_low,

        "time_per_habit": time_per_habit,
        "age_time_interaction": age_time_interaction,
        "habit_time_interaction": habit_time_interaction
    })

    for j in range(Ttr.shape[1]):
        df_feat[f"raw_time_{j+1:02d}"] = Ttr[:, j]

    for j in range(norm_t.shape[1]):
        df_feat[f"norm_time_{j+1:02d}"] = norm_t[:, j]

    for j in range(attempts_tr.shape[1]):
        df_feat[f"attempt_{j+1:02d}"] = attempts_tr[:, j]

    return df_feat.fillna(0)

X_feat = build_participant_features(
    T_raw, age, habit,
    am1_train, amall_train,
    mean_attempt_train, max_attempt_train, fail_rate_train, delayed_rate_train,
    ortho, avg_word_time_train_ref, train_word_idx, codes_pw
)

# ============================================================
# 12. TRAIN / TEST MATRICES
# ============================================================
X_train_full = X_feat.iloc[train_idx].reset_index(drop=True)
X_test_full = X_feat.iloc[test_idx].reset_index(drop=True)

y_train = y_pt_proxy[train_idx]
y_test = y_pt_proxy[test_idx]

id_test = participant_ids[test_idx]

# ============================================================
# 13. TRAIN-ONLY FEATURE SELECTION
# ============================================================
def select_top_features(X_df, y, top_k=40):
    corrs = []
    for col in X_df.columns:
        c = safe_corr(X_df[col].values, y)
        corrs.append(abs(c))
    corrs = np.array(corrs)
    idx = np.argsort(corrs)[-top_k:]
    cols = list(X_df.columns[idx])
    return cols, pd.Series(corrs, index=X_df.columns).sort_values(ascending=False)

TOP_K = 40
selected_cols, corr_series = select_top_features(X_train_full, y_train, top_k=TOP_K)

X_train = X_train_full[selected_cols].copy()
X_test = X_test_full[selected_cols].copy()

print("\nSelected feature matrix shape (train):", X_train.shape)
print("Selected feature matrix shape (test):", X_test.shape)
print("\nTop correlated training features:")
print(corr_series.head(15))

# ============================================================
# 14. VOTING REGRESSOR ONLY
# ============================================================
voting_pipe = Pipeline([
    ("imputer", SimpleImputer(strategy="median")),
    ("model", VotingRegressor([
        ("gb", GradientBoostingRegressor(
            random_state=42,
            n_estimators=500,
            learning_rate=0.02,
            max_depth=2,
            subsample=0.85,
            min_samples_leaf=2,
            loss="huber"
        )),
        ("rf", RandomForestRegressor(
            random_state=42,
            n_estimators=600,
            max_depth=12,
            min_samples_split=2,
            min_samples_leaf=1,
            max_features=0.8,
            n_jobs=-1
        )),
        ("xgb", xgb.XGBRegressor(
            random_state=42,
            objective="reg:squarederror",
            verbosity=0,
            tree_method="hist",
            n_jobs=-1,
            n_estimators=400,
            learning_rate=0.03,
            max_depth=3,
            min_child_weight=3,
            subsample=0.85,
            colsample_bytree=0.85,
            gamma=0.1,
            reg_alpha=0.01,
            reg_lambda=3.0
        ))
    ]))
])

voting_param_grid = {
    "model__weights": [
        (1, 1, 1),
        (2, 1, 1),
        (1, 2, 1),
        (1, 1, 2),
        (2, 1, 2),
        (2, 2, 3),
        (3, 2, 4)
    ]
}

cv = KFold(n_splits=5, shuffle=True, random_state=42)

search = GridSearchCV(
    estimator=voting_pipe,
    param_grid=voting_param_grid,
    scoring="r2",
    cv=cv,
    n_jobs=-1,
    refit=True,
    error_score="raise"
)

print("\n==================== VOTING REGRESSOR ONLY ====================")
search.fit(X_train, y_train)

y_pred = search.best_estimator_.predict(X_test)

mae = mean_absolute_error(y_test, y_pred)
mse = mean_squared_error(y_test, y_pred)
rmse = np.sqrt(mse)
r2 = r2_score(y_test, y_pred)

print("\nBest Params:")
print(search.best_params_)

print("\nMetrics:")
print(f"BestCV_R2 = {search.best_score_:.4f}")
print(f"Test_MAE  = {mae:.4f}")
print(f"Test_MSE  = {mse:.4f}")
print(f"Test_RMSE = {rmse:.4f}")
print(f"Test_R2   = {r2:.4f}")

pred_df = pd.DataFrame({
    "participant_id": id_test,
    "true_avg_time_test6": y_test,
    "pred_avg_time_test6": y_pred
})

print("\nPredictions preview:")
print(pred_df.head())

# Optional save
# pred_df.to_csv("/content/voting_regressor_only_predictions.csv", index=False)
# print("\nSaved: /content/voting_regressor_only_predictions.csv")

import numpy as np
import matplotlib.pyplot as plt
from scipy.stats import linregress
from sklearn.metrics import r2_score, mean_squared_error, mean_absolute_error

# ============================================================
# REGRESSION PLOT: ACTUAL VS PREDICTED
# ============================================================
MODEL_NAME = "xgb_pipe"   # change this to "Lasso", "Ridge", "SVR", etc.

y_true = np.array(y_test, dtype=float)
y_hat = np.array(y_pred, dtype=float)

# metrics
r2 = r2_score(y_true, y_hat)
rmse = np.sqrt(mean_squared_error(y_true, y_hat))
mae = mean_absolute_error(y_true, y_hat)

# axis range
mn = min(y_true.min(), y_hat.min())
mx = max(y_true.max(), y_hat.max())
pad = 0.05 * (mx - mn) if mx > mn else 1.0

xmin, xmax = mn - pad, mx + pad
ymin, ymax = mn - pad, mx + pad

# regression line
slope, intercept, _, _, _ = linregress(y_true, y_hat)
x_line = np.linspace(xmin, xmax, 200)
y_line = slope * x_line + intercept

# plot
plt.figure(figsize=(7, 7), dpi=600)

plt.scatter(
    y_true, y_hat,
    s=40,
    alpha=1,
    color="blue"
)

# ideal diagonal
plt.plot(
    [xmin, xmax], [xmin, xmax],
    linestyle="--",
    linewidth=1.5,
    color="red",
    label="Theoritical Diagonal"
)

# fitted regression line
plt.plot(
    x_line, y_line,
    linewidth=1.5,
    color="blue",
    label=f"Regression line"
)

plt.xlabel("Actual", fontsize=20)
plt.ylabel("Predicted", fontsize=20)
plt.xlim(xmin, xmax)
plt.ylim(ymin, ymax)
plt.legend(fontsize=20)
plt.xticks(fontsize=20)
plt.yticks(fontsize=20)
# plt.grid(alpha=0.3)

# plt.text(
#     0.05, 0.95,
#     f"$R^2$ = {r2:.4f}\nRMSE = {rmse:.4f}\nMAE = {mae:.4f}",
#     transform=plt.gca().transAxes,
#     ha="left", va="top",
#     fontsize=10,
#     bbox=dict(boxstyle="round", facecolor="white", alpha=0.8)
# )

# plt.tight_layout()

# save with regression model name
save_path = f"{MODEL_NAME}_regression_plot.png"
plt.savefig(save_path, dpi=600, bbox_inches="tight")

plt.show()

print("Saved as:", save_path)

